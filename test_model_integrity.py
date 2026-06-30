"""
Model integrity regression tests.

Locks in the correctness work: the 3-statement balance sheet must tie to zero,
the deployment buffer must hold, horizons must be length-safe, and the base-case
headline metrics must stay within tolerance. Runs the engines directly with
defaults (no LLM / no network).

Run:  pytest test_model_integrity.py -q
  or: python test_model_integrity.py
"""
import warnings
warnings.filterwarnings("ignore")

from assumptions.revenue_defaults import get_default_revenue_assumptions
from assumptions.opex_defaults import get_default_opex_assumptions
from assumptions.capex_defaults import get_default_capex_assumptions
from assumptions.depreciation_defaults import get_default_depreciation_assumptions
from assumptions.loan_defaults import get_default_loan_assumptions
from assumptions.tax_defaults import get_default_tax_assumptions
from assumptions.working_capital_defaults import get_default_working_capital_assumptions
from assumptions.valuation_defaults import get_default_valuation_assumptions

from src.engines.revenue_engine import compute_revenue
from src.engines.opex_engine import compute_opex
from src.engines.capex_engine import compute_capex
from src.engines.depreciation_engine import compute_depreciation
from src.engines.loan_engine import compute_loan
from src.engines.tax_engine import compute_tax
from src.engines.working_capital_engine import compute_working_capital
from src.engines.cashflow_engine import compute_cashflow

BASE = {
    "location": "Mumbai", "total_racks": 1000, "facility_type": "retail_colo",
    "projection_years": 10, "start_year": 2026,
}

SCENARIOS = [
    {"location": "Hyderabad", "total_racks": 1500, "facility_type": "wholesale",   "projection_years": 15, "start_year": 2026},
    {"location": "Bangalore", "total_racks": 500,  "facility_type": "retail_colo", "projection_years": 20, "start_year": 2026},
    {"location": "Delhi",     "total_racks": 2000, "facility_type": "ai_hpc",      "projection_years": 12, "start_year": 2026},
]


def run_model(user_inputs):
    """Run the full pipeline with default assumptions (no LLM). CapEx owns the
    deployment schedule; revenue occupancy is capped by it (single source)."""
    cap = compute_capex(user_inputs, get_default_capex_assumptions())
    ui = {**user_inputs, "deployment_schedule": {
        p["year"]: p["racks"] for p in cap["deployment_schedule"]
    }}
    rev = compute_revenue(ui, get_default_revenue_assumptions())
    opx = compute_opex(rev, cap, get_default_opex_assumptions())
    dep = compute_depreciation(cap, get_default_depreciation_assumptions())
    loan = compute_loan(cap, get_default_loan_assumptions())
    tax = compute_tax(opx, dep, loan, get_default_tax_assumptions())
    wc = compute_working_capital(rev, get_default_working_capital_assumptions(), opx)
    cf = compute_cashflow(opx, cap, dep, loan, tax, wc, get_default_valuation_assumptions())
    return dict(rev=rev, cap=cap, opx=opx, dep=dep, loan=loan, tax=tax, wc=wc, cf=cf)


# ── Balance sheet integrity ─────────────────────────────────────────────────

def test_balance_sheet_ties_base():
    cf = run_model(BASE)["cf"]
    for yr, chk in zip(cf["metadata"]["years"], cf["balance_sheet"]["balance_check"]):
        assert abs(chk) < 1e-6, f"balance check {chk} != 0 in {yr}"


def test_balance_sheet_ties_all_scenarios():
    for s in SCENARIOS:
        cf = run_model(s)["cf"]
        for chk in cf["balance_sheet"]["balance_check"]:
            assert abs(chk) < 1e-6, f"balance broke in {s['location']}/{s['facility_type']}"


# ── Capacity coherence (single source of truth) ─────────────────────────────

def test_occupancy_never_exceeds_deployment():
    for s in [BASE] + SCENARIOS:
        out = run_model(s)
        occ = out["rev"]["drivers"]["occupied_racks"]
        rd = out["cap"]["drivers"]["racks_deployed"]
        cum, running = [], 0
        for x in rd:
            running += x
            cum.append(running)
        for i in range(len(occ)):
            assert cum[i] >= occ[i] - 0.01, f"occupancy exceeds deployment in {s['location']}"


# ── Dynamic horizon length-safety ───────────────────────────────────────────

def test_horizons_length_safe():
    for yrs in (5, 10, 15, 20):
        out = run_model({**BASE, "projection_years": yrs})
        assert len(out["cf"]["metadata"]["years"]) == yrs
        assert len(out["rev"]["drivers"]["occupied_racks"]) == yrs
        for chk in out["cf"]["balance_sheet"]["balance_check"]:
            assert abs(chk) < 1e-6


# ── Construction year discipline ────────────────────────────────────────────

def test_construction_year_is_clean():
    out = run_model(BASE)
    # Year 0 (construction): zero revenue, zero opex, zero depreciation
    assert out["rev"]["revenue_streams"]["net_revenue"][0] == 0.0
    assert out["opx"]["financials"]["total_opex"][0] == 0.0
    assert out["dep"]["financials"]["book_depreciation"][0] == 0.0


# ── Base-case headline metrics (regression guard) ───────────────────────────

def test_base_case_metrics_within_tolerance():
    cf = run_model(BASE)["cf"]
    v = cf["valuation"]
    assert abs(v["npv"] - 124.8) < 2.0,                f"NPV drifted: {v['npv']}"
    assert abs(v["project_irr"] - 0.185) < 0.01,       f"Project IRR drifted: {v['project_irr']}"
    assert abs(v["equity_irr"] - 0.216) < 0.01,        f"Equity IRR drifted: {v['equity_irr']}"
    assert abs(cf["equity"]["moic"] - 4.74) < 0.3,     f"MOIC drifted: {cf['equity']['moic']}"


def test_dscr_profile_shape():
    cf = run_model(BASE)["cf"]
    dscr = [d for d in cf["cashflows"]["dscr"] if d is not None]
    # Greenfield shape: ramps up, ends well above covenant
    assert dscr[1] < 0.5            # early lease-up year is weak
    assert dscr[-1] > 2.0           # stabilized years comfortably covered
    # non-construction DSCR should be non-decreasing in the back half
    assert dscr[-1] >= dscr[-3]


# ── Excel generation (no crashes across scenarios/horizons) ─────────────────

def _excel_override(user_inputs):
    """Build a generate() override using default assumptions (no LLM)."""
    return {
        "ui":   user_inputs,
        "rev":  get_default_revenue_assumptions(),
        "cap":  get_default_capex_assumptions(),
        "opx":  get_default_opex_assumptions(),
        "dep":  get_default_depreciation_assumptions(),
        "loan": get_default_loan_assumptions(),
        "tax":  get_default_tax_assumptions(),
        "wc":   get_default_working_capital_assumptions(),
        "val":  get_default_valuation_assumptions(),
    }


def test_excel_generates_for_all_scenarios():
    import tempfile, os
    from src.reporting.excel_generator import generate
    from openpyxl import load_workbook
    cases = [BASE] + SCENARIOS + [{**BASE, "projection_years": y} for y in (5, 20)]
    for s in cases:
        fd, path = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
        try:
            generate(path, override=_excel_override(s))
            wb = load_workbook(path)
            assert "BS" in wb.sheetnames, f"BS tab missing for {s}"
        finally:
            if os.path.exists(path):
                os.remove(path)


# ── Excel ↔ engine parity (skips if the formula evaluator isn't installed) ──

def _assert_excel_matches_engine(user_inputs):
    """Generate the workbook for `user_inputs`, evaluate its formulas, and assert
    the headline lines match the engine and the balance sheet ties to zero."""
    import tempfile, os
    import openpyxl.utils as U
    from openpyxl import load_workbook
    import formulas
    from src.reporting.excel_generator import generate

    n = user_inputs["projection_years"]
    fd, path = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
    try:
        generate(path, override=_excel_override(user_inputs))
        wb = load_workbook(path)
        sol = formulas.ExcelModel().loads(path).finish().calculate()
        fname = os.path.basename(path)

        def label_row(sheet, prefix):
            for row in wb[sheet].iter_rows(min_col=1, max_col=2):
                for c in row:
                    if isinstance(c.value, str) and c.value.strip().startswith(prefix):
                        return c.row
            return None

        def series(sheet, row):
            out = []
            for j in range(n):
                col = U.get_column_letter(3 + j)
                v = sol.get("'[%s]%s'!%s%d" % (fname, sheet, col, row))
                try:
                    out.append(float(v.value[0, 0]))
                except Exception:
                    out.append(0.0)
            return out

        eng = run_model(user_inputs)
        checks = [
            ("CAPEX", "Total CapEx",             eng["cap"]["financials"]["total_capex"]),
            ("DEBT",  "Total interest expense",  eng["loan"]["long_term_debt_account"]["interest_expense"]),
            ("DEBT",  "Closing debt",            eng["loan"]["long_term_debt_account"]["closing_balance"]),
            ("PNL",   "EBITDA",                  eng["cf"]["pnl"]["ebitda"]),
            ("TAX",   "Income tax payable",      eng["tax"]["financials"]["tax"]),
            ("CFS",   "CFADS",                   eng["cf"]["cashflows"]["cfads"]),
        ]
        for sheet, prefix, eng_vals in checks:
            xl_vals = series(sheet, label_row(sheet, prefix))
            for i, (a, b) in enumerate(zip(xl_vals, eng_vals[:n])):
                assert abs(a - b) < 0.6, f"[{user_inputs['location']}/{user_inputs['total_racks']}r] {sheet}/{prefix} yr{i}: excel {a:.2f} vs engine {b:.2f}"

        for v in series("BS", label_row("BS", "Balance check")):
            assert abs(v) < 1e-3, "Excel balance sheet does not tie to zero"

        # Valuation scalars (NPV / IRR / MOIC) — the returns block, where the
        # terminal-value and horizon bugs hid. These must match the engine.
        def val_scalar(prefix):
            for row in wb["VAL"].iter_rows(min_col=1, max_col=2):
                for c in row:
                    if isinstance(c.value, str) and c.value.strip().startswith(prefix):
                        v = sol.get("'[%s]VAL'!C%d" % (fname, c.row))
                        try:
                            return float(v.value[0, 0])
                        except Exception:
                            return None
            return None

        ev = eng["cf"]["valuation"]
        tag = f"[{user_inputs['location']}/{user_inputs['total_racks']}r/{n}y]"
        assert abs(val_scalar("NPV") - ev["npv"]) < 2.0, f"{tag} NPV: excel {val_scalar('NPV'):.1f} vs engine {ev['npv']:.1f}"
        assert abs(val_scalar("Project IRR") - ev["project_irr"]) < 0.006, f"{tag} Project IRR mismatch"
        assert abs(val_scalar("Equity IRR") - ev["equity_irr"]) < 0.006, f"{tag} Equity IRR mismatch"
        assert abs(val_scalar("MOIC") - eng["cf"]["equity"]["moic"]) < 0.15, f"{tag} MOIC mismatch"
    finally:
        if os.path.exists(path):
            os.remove(path)


def test_excel_matches_engine_base():
    try:
        import formulas  # noqa: F401
    except ImportError:
        import pytest
        pytest.skip("`formulas` not installed — parity check skipped")
    _assert_excel_matches_engine(BASE)


def test_excel_matches_engine_nondefault():
    # Non-default rack count + location + facility AND a 15-year horizon — the
    # 15-year case exercises the lease-up padding + dynamic valuation ranges that
    # broke the Excel returns block beyond year 10.
    try:
        import formulas  # noqa: F401
    except ImportError:
        import pytest
        pytest.skip("`formulas` not installed — parity check skipped")
    _assert_excel_matches_engine({
        "location": "Hyderabad", "total_racks": 1500, "facility_type": "wholesale",
        "projection_years": 15, "start_year": 2026,
    })


if __name__ == "__main__":
    import traceback
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    passed = 0
    for t in tests:
        try:
            t()
            print(f"PASS  {t.__name__}")
            passed += 1
        except Exception:
            print(f"FAIL  {t.__name__}")
            traceback.print_exc()
    print(f"\n{passed}/{len(tests)} passed")
