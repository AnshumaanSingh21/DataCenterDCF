"""
Formula-driven Excel DCF workbook generator.
Run:  python -m src.reporting.excel_generator
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from assumptions.revenue_defaults       import get_default_revenue_assumptions, FACILITY_TYPE_OVERRIDES
from assumptions.capex_defaults         import get_default_capex_assumptions
from assumptions.opex_defaults          import get_default_opex_assumptions
from assumptions.depreciation_defaults  import get_default_depreciation_assumptions
from assumptions.loan_defaults          import get_default_loan_assumptions
from assumptions.tax_defaults           import get_default_tax_assumptions
from assumptions.working_capital_defaults import get_default_working_capital_assumptions
from assumptions.valuation_defaults     import get_default_valuation_assumptions
from src.engines.revenue_engine         import compute_revenue
from src.engines.capex_engine           import compute_capex
from src.engines.opex_engine            import compute_opex
from src.engines.depreciation_engine    import compute_depreciation
from src.engines.loan_engine            import compute_loan
from src.engines.tax_engine             import compute_tax
from src.engines.working_capital_engine import compute_working_capital
from src.engines.cashflow_engine        import compute_cashflow

# ─── constants ────────────────────────────────────────────────────────────────
USER_INPUTS = {
    "location": "Mumbai", "total_racks": 1000, "facility_type": "retail_colo",
    "projection_years": 10, "start_year": 2026,
    # deployment_schedule is derived from CapEx (single source of truth) in _run_pipeline
}
_DEFAULT_USER_INPUTS = dict(USER_INPUTS)  # immutable snapshot for reset
N = USER_INPUTS["projection_years"]
YEARS = list(range(USER_INPUTS["start_year"], USER_INPUTS["start_year"] + N))

# When the workbook is generated for a specific user run, generate() injects
# the run's inputs + assumption dicts here so the Excel matches the dashboard
# (location, racks, facility type, horizon, and all LLM/UI-overridden values).
# When None, the generator falls back to the module defaults above.
_OVERRIDE = None

def _A(key, default_fn):
    """Return the overridden assumption dict for `key` if a run was injected,
    else the engine default."""
    if _OVERRIDE and _OVERRIDE.get(key) is not None:
        return _OVERRIDE[key]
    return default_fn()
COL_LBL = 1; COL_UNIT = 2; COL_YR0 = 3
DEPLOY_PY   = [0, 3, 6]
DEPLOY_XCOL = ['C', 'F', 'I']
DEPLOY_XIDX = [1, 4, 7]

# ─── colours ──────────────────────────────────────────────────────────────────
NAVY    = "1F3864"; DKGREY  = "404040"; MDGREY  = "808080"
LTGREY  = "F2F2F2"; BLUE_IN = "EBF3FB"; WHITE   = "FFFFFF"
GRNCELL = "E8F5E9"; REDDARK = "C00000"; SRC_IN  = "DAEEF3"  # market-sourced (AI) input — light teal

# ─── style helpers ────────────────────────────────────────────────────────────
def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, col="1A1A1A", sz=10, ital=False):
    return Font(name="Calibri", bold=bold, color=col, size=sz, italic=ital)
def _aln(h="right", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _nfmt(c, fmt):
    """Set number_format only if fmt is not None (guards against openpyxl TypeError)."""
    if fmt is not None:
        c.number_format = fmt

def _bdr(top=False, bot=False):
    t = Side(style="thin", color="AAAAAA") if top else None
    b = Side(style="thin", color="AAAAAA") if bot else None
    return Border(top=t, bottom=b)

FMT_CR  = '#,##0.00;[Red]-#,##0.00'
FMT_CR0 = '#,##0;[Red]-#,##0'
FMT_P1  = '0.0%'
FMT_P2  = '0.00%'
FMT_MX  = '0.00"x"'
FMT_INT = '#,##0'
FMT_YR  = '0'

def _w(ws, row, col, val=None, fmt=None, bold=False, fill=None,
       color="1A1A1A", ah="right", ital=False, bdr=False):
    c = ws.cell(row=row, column=col)
    if val is not None: c.value = val
    c.font      = _font(bold, color, ital=ital)
    c.alignment = _aln(h=ah, v="center")
    if fill:  c.fill   = _fill(fill)
    if fmt:   c.number_format = fmt
    if bdr:   c.border = _bdr(top=True, bot=True)
    return c

def _hdr(ws, row, title, nc=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1)
    c.value = f"  {title}"; c.font = _font(True, WHITE)
    c.fill = _fill(NAVY); c.alignment = _aln("left")

def _subhdr(ws, row, title, nc=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1)
    c.value = f"  {title}"; c.font = _font(True, DKGREY)
    c.fill = _fill("D9D9D9"); c.alignment = _aln("left")

def _lbl(ws, row, label, unit="", bold=False, italic=False, fill=None):
    lc = ws.cell(row=row, column=COL_LBL)
    lc.value = label; lc.font = _font(bold, ital=italic)
    lc.alignment = _aln("left")
    if fill: lc.fill = _fill(fill)
    uc = ws.cell(row=row, column=COL_UNIT)
    uc.value = unit; uc.font = _font(col=MDGREY, sz=9)
    uc.alignment = _aln("left")
    if fill: uc.fill = _fill(fill)

def _title_row(ws, title, sub=""):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_YR0+N)
    c = ws.cell(row=1, column=1)
    c.value = f"  {title}"; c.font = _font(True, WHITE, 12)
    c.fill = _fill(NAVY); c.alignment = _aln("left"); ws.row_dimensions[1].height = 22
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=COL_YR0+N)
        s = ws.cell(row=2, column=1)
        s.value = f"  {sub}"; s.font = _font(col=MDGREY, ital=True)
        s.alignment = _aln("left")

def _yr_hdrs(ws, r_lbl, r_idx):
    for j, yr in enumerate(YEARS):
        col = COL_YR0 + j
        _w(ws, r_lbl, col, yr,  FMT_YR, bold=True, fill=LTGREY, ah="center")
        _w(ws, r_idx, col, j+1, FMT_YR, color=MDGREY, ah="center")
    _w(ws, r_lbl, COL_YR0+N, "Total", bold=True, fill=LTGREY, ah="center")

def _col_widths(ws):
    ws.column_dimensions[get_column_letter(COL_LBL)].width  = 34
    ws.column_dimensions[get_column_letter(COL_UNIT)].width = 10
    for j in range(N): ws.column_dimensions[get_column_letter(COL_YR0+j)].width = 11
    ws.column_dimensions[get_column_letter(COL_YR0+N)].width = 14

def _sum_row(ws, row, data_rows, fill=LTGREY):
    for j in range(N):
        c = get_column_letter(COL_YR0+j)
        terms = "+".join(f"{c}{r}" for r in data_rows)
        ws.cell(row=row, column=COL_YR0+j).value = f"={terms}"
        _w(ws, row, COL_YR0+j, fmt=FMT_CR, bold=True, fill=fill, bdr=True)
    _w(ws, row, COL_LBL,  fill=fill, bdr=True)
    _w(ws, row, COL_UNIT, fill=fill, bdr=True)

def cl(j):
    """Column letter for year index j (0-based). j=0 → 'C'."""
    return get_column_letter(COL_YR0 + j)

# ASMP row registry — must be filled before any other sheet writer
AR = {}

def _asmp(name):
    """Absolute ref to ASMP scalar (col C)."""
    return f"ASMP!$C${AR[name]}"

def _asmp_arr(name, j):
    """Ref to ASMP array row at year column j."""
    return f"ASMP!{cl(j)}${AR[name]}"

# ─── sheet row registries ─────────────────────────────────────────────────────
SIZE_R = {}; REV_R  = {}; CAP_R  = {}; OPX_R  = {}
DEP_R  = {}; DBT_R  = {}; TAX_R  = {}; WC_R   = {}
PNL_R  = {}; CFS_R  = {}; VAL_R  = {}; BS_R   = {}

# ─── pipeline run ─────────────────────────────────────────────────────────────
def _run_pipeline():
    # CapEx first — it owns the deployment schedule; revenue occupancy
    # is capped by what CapEx actually fitted out (single source of truth).
    cap = compute_capex(USER_INPUTS, _A("cap", get_default_capex_assumptions))
    ui  = {**USER_INPUTS, "deployment_schedule": {
        p["year"]: p["racks"] for p in cap["deployment_schedule"]
    }}
    rev = compute_revenue(ui, _A("rev", get_default_revenue_assumptions))
    opx = compute_opex(rev, cap, _A("opx", get_default_opex_assumptions))
    dep = compute_depreciation(cap, _A("dep", get_default_depreciation_assumptions))
    loan = compute_loan(cap, _A("loan", get_default_loan_assumptions))
    tax = compute_tax(opx, dep, loan, _A("tax", get_default_tax_assumptions))
    wc  = compute_working_capital(rev, _A("wc", get_default_working_capital_assumptions), opx)
    cf  = compute_cashflow(opx, cap, dep, loan, tax, wc, _A("val", get_default_valuation_assumptions))
    return dict(rev=rev, cap=cap, opx=opx, dep=dep, loan=loan, tax=tax, wc=wc, cf=cf)

# ═══════════════════════════════════════════════════════════════════════════════
# ASMP — all hardcoded inputs
# ═══════════════════════════════════════════════════════════════════════════════
def write_asmp(wb, P):
    ws = wb.create_sheet("ASMP")
    _col_widths(ws)
    _title_row(ws, "ASSUMPTIONS", "All inputs — change here only")
    ws.freeze_panes = "C5"

    cc = P['cap']['capex_components']
    rev_a = dict(_A("rev", get_default_revenue_assumptions))
    rev_a.update(FACILITY_TYPE_OVERRIDES.get(USER_INPUTS.get("facility_type", "retail_colo"), {}))
    loan_a = _A("loan", get_default_loan_assumptions)
    cap_a  = _A("cap", get_default_capex_assumptions)
    opx_a  = _A("opx", get_default_opex_assumptions)
    dep_a  = _A("dep", get_default_depreciation_assumptions)
    tax_a  = _A("tax", get_default_tax_assumptions)
    wc_a   = _A("wc", get_default_working_capital_assumptions)
    val_a  = _A("val", get_default_valuation_assumptions)

    lease_up   = rev_a['lease_up_curve']

    r = 3  # row counter (rows 1–2 = title/subtitle)
    _yr_hdrs(ws, r, r+1); r += 2  # rows 3,4

    def inp(row, label, unit, val, fmt=FMT_CR, col=None, ah="right", src=False):
        # src=True → market-sourced (AI/market-intelligence) value, coloured
        # distinctly so it isn't mistaken for a hand-entered hardcode.
        _lbl(ws, row, label, unit)
        c = ws.cell(row=row, column=COL_YR0)
        c.value = val; c.fill = _fill(SRC_IN if src else BLUE_IN); c.font = _font(col=col or "1F3864", bold=True)
        c.alignment = _aln(h=ah, v="center")
        if fmt: c.number_format = fmt

    def arr(row, label, unit, vals, fmt=FMT_CR):
        _lbl(ws, row, label, unit)
        for j, v in enumerate(vals):
            c = ws.cell(row=row, column=COL_YR0+j)
            c.value = v; c.fill = _fill(BLUE_IN); c.font = _font(col="1F3864", bold=True)
            c.number_format = fmt; c.alignment = _aln(v="center")

    def der(row, label, unit, formula, fmt=FMT_CR):
        """Derived (formula) row in ASMP."""
        _lbl(ws, row, label, unit, italic=True)
        c = ws.cell(row=row, column=COL_YR0)
        c.value = formula; c.fill = _fill("FFF9E6"); c.font = _font(col=DKGREY, ital=True)
        c.number_format = fmt; c.alignment = _aln(v="center")

    # ── PROJECT OVERVIEW ──────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "PROJECT OVERVIEW"); r += 1
    AR['total_racks'] = r;    inp(r, "Total Racks",         "racks", 1000,              FMT_INT); r += 1
    AR['location']    = r;    inp(r, "Location",             "–",    "Mumbai",           None, ah="left"); r += 1
    AR['facility_type']= r;   inp(r, "Facility Type",        "–",    "retail_colo",      None, ah="left"); r += 1
    AR['start_year']  = r;    inp(r, "Start Year",           "–",    2026,               FMT_YR); r += 1
    AR['proj_years']  = r;    inp(r, "Projection Years",     "yrs",  10,                 FMT_INT); r += 1
    AR['facility_sqft']= r;   inp(r, "Facility Floor Area",  "sqft", 100000,             FMT_INT); r += 1

    r += 1; _hdr(ws, r, "DEPLOYMENT SCHEDULE"); r += 1
    AR['phase1_pct'] = r; inp(r, "Phase 1 — % of total racks (Year 1)", "%",   cap_a['phase_1_pct'], FMT_P1); r += 1
    AR['phase2_pct'] = r; inp(r, "Phase 2 — % of total racks (Year 4)", "%",   cap_a['phase_2_pct'], FMT_P1); r += 1
    AR['phase3_pct'] = r; inp(r, "Phase 3 — % of total racks (Year 7)", "%",   cap_a['phase_3_pct'], FMT_P1); r += 1

    # Derived: rack counts from percentages × total_racks
    AR['deploy_arr'] = r
    _lbl(ws, r, "Racks deployed (this year, derived)", "racks", italic=True)
    _phase_pct_map = {0: 'phase1_pct', 3: 'phase2_pct', 6: 'phase3_pct'}
    for j in range(N):
        c = ws.cell(row=r, column=COL_YR0+j)
        if j in _phase_pct_map:
            c.value = f"=ROUND({_asmp('total_racks')}*ASMP!$C${AR[_phase_pct_map[j]]},0)"
        else:
            c.value = 0
        c.fill = _fill("FFF9E6"); c.font = _font(col=DKGREY, ital=True)
        c.number_format = FMT_INT; c.alignment = _aln(v="center")
    r += 1

    r += 1; _hdr(ws, r, "LEASE-UP CURVE"); r += 1
    AR['lease_up']    = r;    arr(r, "Lease-up (% of total racks)",  "%", lease_up, FMT_P1); r += 1

    # ── REVENUE ───────────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "REVENUE ASSUMPTIONS"); r += 1
    AR['rack_mrc']      = r; inp(r, "Rack MRC (Year 1)",            "Cr/rack/mo",  rev_a['rack_mrc_crore'],             FMT_CR, src=True); r += 1
    AR['rack_mrc_esc']  = r; inp(r, "Rack MRC escalation",          "% p.a.",      rev_a.get('rack_mrc_escalation', 0.05), FMT_P1); r += 1
    AR['otc_fee']       = r; inp(r, "OTC fee per new rack (Yr 1)",  "Cr/rack",     rev_a.get('otc_fee_crore', 0.0003),    FMT_CR); r += 1
    AR['otc_esc']       = r; inp(r, "OTC fee escalation",           "% p.a.",      0.05,     FMT_P1); r += 1
    AR['util_tariff']   = r; inp(r, "Grid tariff (Year 1)",         "Rs/kWh",      rev_a['utility_tariff_rs_per_kwh'],    FMT_CR, src=True); r += 1
    AR['pwr_markup']    = r; inp(r, "Power markup",                 "Rs/kWh",      rev_a['power_markup_rs_per_kwh'],      FMT_CR, src=True); r += 1
    AR['tenant_tariff'] = r
    der(r, "Tenant power tariff (Year 1)", "Rs/kWh", f"=ASMP!$C${r-2}+ASMP!$C${r-1}", FMT_CR); r += 1
    AR['pwr_esc']       = r; inp(r, "Power tariff escalation",      "% p.a.",      0.05,     FMT_P1); r += 1
    AR['pue']           = r; inp(r, "PUE",                          "–",           rev_a['pue'],  FMT_CR, src=True); r += 1
    AR['kw_per_rack']   = r; inp(r, "IT load per rack",             "kW/rack",     rev_a.get('kw_per_rack', 6.0), FMT_CR); r += 1
    AR['dot_share']     = r; inp(r, "DoT revenue share",            "%",           0.0,      FMT_P2); r += 1

    # ── CAPEX ─────────────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "CAPEX ASSUMPTIONS"); r += 1
    # Per-rack fit-out cost = year-0 component capex ÷ actual phase-1 racks
    _ph1 = P['cap']['drivers']['racks_deployed'][0] or 1
    net_per_rack  = round(cc['network_capex'][0]    / _ph1, 6)
    elec_per_rack = round(cc['electrical_capex'][0] / _ph1, 6)
    AR['civil_pr']   = r; inp(r, "Civil cost per rack",          "Cr/rack",  cap_a['civil_cost_per_rack'],       FMT_CR, src=True); r += 1
    AR['elec_pr']    = r; inp(r, "Electrical cost per rack",     "Cr/rack",  elec_per_rack,                     FMT_CR, src=True); r += 1
    AR['mech_pr']    = r; inp(r, "Mechanical cost per rack",     "Cr/rack",  cap_a['mechanical_cost_per_rack'],  FMT_CR, src=True); r += 1
    AR['it_pr']      = r; inp(r, "IT hardware cost per rack",    "Cr/rack",  cap_a['it_hardware_cost_per_rack'], FMT_CR); r += 1
    AR['net_pr']     = r; inp(r, "Network cost per rack",        "Cr/rack",  net_per_rack,  FMT_CR); r += 1
    AR['preop_pct']  = r; inp(r, "Pre-operational (% hard cost)","% hard",   cap_a.get('pre_op_pct', 0.10), FMT_P2); r += 1
    AR['software_c'] = r; inp(r, "Software CapEx (Phase 1 only)","Cr",       10.0,          FMT_CR); r += 1
    AR['land_c']       = r; inp(r, "Land cost (Phase 1 only)",              "Cr", P['cap']['site_sizing']['land_cost_crore'], FMT_CR, src=True); r += 1
    AR['siteprep_c']   = r; inp(r, "Site prep (consult + approvals, Ph1)", "Cr", 15.0,                                       FMT_CR); r += 1
    AR['misc_infra_c'] = r; inp(r, "Misc infrastructure (HT line, facade, commissioning, Ph1)", "Cr", cap_a.get('misc_infrastructure_cost_crore', 0.0), FMT_CR); r += 1

    # ── OPEX ──────────────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "OPEX ASSUMPTIONS"); r += 1
    AR['fte_per100'] = r; inp(r, "FTE per 100 racks",            "FTE",      3,             FMT_INT); r += 1
    AR['avg_ctc']    = r; inp(r, "Avg CTC per employee",         "Rs lakh",  12,            FMT_CR); r += 1
    AR['mp_esc']     = r; inp(r, "Manpower escalation",          "% p.a.",   0.08,          FMT_P1); r += 1
    AR['hk_rate']    = r; inp(r, "Housekeeping rate",            "Rs/sqft/yr",150,          FMT_INT); r += 1
    AR['hk_esc']     = r; inp(r, "Housekeeping escalation",      "% p.a.",   0.05,          FMT_P1); r += 1
    AR['amc_civil']  = r; inp(r, "AMC — civil",                  "% of cumul civil",  0.02, FMT_P2); r += 1
    AR['amc_elec']   = r; inp(r, "AMC — electrical",             "% of cumul elec",   0.04, FMT_P2); r += 1
    AR['amc_mech']   = r; inp(r, "AMC — mechanical",             "% of cumul mech",   0.08, FMT_P2); r += 1
    AR['amc_net']    = r; inp(r, "AMC — network",                "% of cumul net",    0.08, FMT_P2); r += 1
    AR['amc_soft']   = r; inp(r, "AMC — software",               "% of cumul soft",   0.12, FMT_P2); r += 1
    AR['net_pct']    = r; inp(r, "Network opex (% of revenue)",  "%",        0.02,          FMT_P2); r += 1
    AR['sec_pct']    = r; inp(r, "Security (% of revenue)",      "%",        0.01,          FMT_P2); r += 1
    AR['ins_pct']    = r; inp(r, "Insurance (% of cumul CapEx)", "%",        0.005,         FMT_P2); r += 1
    AR['ptax_pct']   = r; inp(r, "Property tax (% of land + civil)","%",     opx_a.get('property_tax_pct_of_asset_value', 0.0022), FMT_P2); r += 1
    AR['gna_pct']    = r; inp(r, "G&A (% of net revenue)",       "%",        0.03,          FMT_P2); r += 1
    AR['mkt_pct_start']  = r; inp(r, "Marketing — Year 1 (% of net revenue)",  "%", 0.01,   FMT_P2); r += 1
    AR['mkt_pct_end']    = r; inp(r, "Marketing — Year 10 (% of net revenue)", "%", 0.0025, FMT_P2); r += 1
    AR['maint_cx']           = r; inp(r, "Maintenance CapEx rate (% of MEP)",       "%", 0.015,  FMT_P2);  r += 1
    AR['maint_warranty']     = r; inp(r, "OEM warranty period",                    "yrs", 3,     FMT_INT); r += 1
    AR['construction_years'] = r; inp(r, "Construction period",                    "yrs", 1,     FMT_INT); r += 1

    # ── DEPRECIATION ──────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "DEPRECIATION — USEFUL LIVES (SLM)"); r += 1
    AR['life_civil'] = r; inp(r, "Civil & structural",           "years",    30,  FMT_INT); r += 1
    AR['life_elec']  = r; inp(r, "Electrical systems",           "years",    15,  FMT_INT); r += 1
    AR['life_mech']  = r; inp(r, "Mechanical systems",           "years",    15,  FMT_INT); r += 1
    AR['life_it']    = r; inp(r, "IT hardware",                  "years",    5,   FMT_INT); r += 1
    AR['life_net']   = r; inp(r, "Network infrastructure",       "years",    7,   FMT_INT); r += 1
    AR['life_soft']  = r; inp(r, "Software & licensing",         "years",    5,   FMT_INT); r += 1

    # ── FINANCING ─────────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "FINANCING ASSUMPTIONS"); r += 1
    AR['debt_pct']   = r; inp(r, "Debt / Total CapEx",           "%",        loan_a['debt_pct'], FMT_P2); r += 1
    AR['eq_pct']     = r
    der(r, "Equity / Total CapEx", "%", f"=1-ASMP!$C${r-1}", FMT_P2); r += 1
    AR['int_rate']   = r; inp(r, "Interest rate",                "% p.a.",   loan_a['interest_rate'], FMT_P2, src=True); r += 1
    AR['morat']      = r; inp(r, "Moratorium",                   "years",    loan_a['moratorium_years'], FMT_INT); r += 1
    AR['tenure']     = r; inp(r, "Loan tenure",                  "years",    10,   FMT_INT); r += 1

    # ── TAX ───────────────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "TAX"); r += 1
    AR['tax_rate']   = r; inp(r, "Corporate tax rate",           "%",        0.25, FMT_P2); r += 1

    # ── WORKING CAPITAL ───────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "WORKING CAPITAL"); r += 1
    AR['dso']        = r; inp(r, "Receivable days (DSO)",         "days",   wc_a.get('receivable_days', 30), FMT_INT); r += 1
    AR['dpo']        = r; inp(r, "Payable days (DPO)",            "days",   wc_a.get('payable_days', 30),    FMT_INT); r += 1

    # ── VALUATION ─────────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "VALUATION"); r += 1
    AR['cost_eq']    = r; inp(r, "Cost of equity",               "% p.a.",   0.18, FMT_P2); r += 1
    AR['ev_mult']    = r; inp(r, "Terminal EV/EBITDA multiple",  "x",        12.0, FMT_MX); r += 1

    _col_widths(ws)
    ws.row_dimensions[3].height = 18
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SIZE — Sizing & occupancy
# ═══════════════════════════════════════════════════════════════════════════════
def write_size(wb):
    ws = wb.create_sheet("SIZE")
    _col_widths(ws); _title_row(ws, "SIZING", "Rack deployment, occupancy & power")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, formula, fmt=FMT_CR, fill=None):
        ws.cell(row=row, column=COL_YR0+j).value = formula
        ws.cell(row=row, column=COL_YR0+j).number_format = fmt
        if fill: ws.cell(row=row, column=COL_YR0+j).fill = _fill(fill)

    # ── Deployment ────────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "DEPLOYMENT"); r += 1
    SIZE_R['deploy'] = r
    _lbl(ws, r, "Racks deployed (this year)", "racks")
    for j in range(N): f(r, j, f"=ASMP!{cl(j)}${AR['deploy_arr']}", FMT_INT)
    SIZE_R['cumul'] = r + 1
    _lbl(ws, r+1, "Cumulative racks deployed", "racks", bold=True)
    for j in range(N):
        if j == 0:
            f(r+1, j, f"=SIZE!{cl(0)}{r}", FMT_INT, LTGREY)
        else:
            f(r+1, j, f"=SIZE!{cl(j-1)}{r+1}+SIZE!{cl(j)}{r}", FMT_INT, LTGREY)
    r += 2

    # ── Occupancy ─────────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "OCCUPANCY"); r += 1
    SIZE_R['lease_up'] = r
    _lbl(ws, r, "Lease-up rate (% of total racks)", "%")
    for j in range(N): f(r, j, f"=ASMP!{cl(j)}${AR['lease_up']}", FMT_P1)
    r += 1

    SIZE_R['gross_occ'] = r
    _lbl(ws, r, "Gross occupied racks (lease-up × total)", "racks")
    for j in range(N):
        f(r, j, f"={_asmp_arr('lease_up',j)}*{_asmp('total_racks')}", FMT_CR)
    r += 1

    SIZE_R['occupied'] = r
    _lbl(ws, r, "Occupied racks (capped at deployed)", "racks", bold=True)
    for j in range(N):
        g = SIZE_R['gross_occ']; d = SIZE_R['cumul']
        f(r, j, f"=MIN(SIZE!{cl(j)}{g},SIZE!{cl(j)}{d})", FMT_CR, LTGREY)
    r += 1

    SIZE_R['new_racks'] = r
    _lbl(ws, r, "New racks taken up this year", "racks")
    for j in range(N):
        if j == 0:
            f(r, j, f"=SIZE!{cl(0)}{SIZE_R['occupied']}", FMT_CR)
        else:
            o = SIZE_R['occupied']
            f(r, j, f"=MAX(SIZE!{cl(j)}{o}-SIZE!{cl(j-1)}{o},0)", FMT_CR)
    r += 2

    # ── Power ─────────────────────────────────────────────────────────────
    _hdr(ws, r, "POWER"); r += 1
    SIZE_R['it_load'] = r
    _lbl(ws, r, "IT load", "kW")
    for j in range(N):
        f(r, j, f"=SIZE!{cl(j)}{SIZE_R['occupied']}*{_asmp('kw_per_rack')}", FMT_CR)
    r += 1

    SIZE_R['fac_load'] = r
    _lbl(ws, r, "Facility load (IT × PUE)", "kW", bold=True)
    for j in range(N):
        f(r, j, f"=SIZE!{cl(j)}{SIZE_R['it_load']}*{_asmp('pue')}", FMT_CR, LTGREY)
    r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# REV — Revenue build
# ═══════════════════════════════════════════════════════════════════════════════
def write_rev(wb):
    ws = wb.create_sheet("REV")
    _col_widths(ws); _title_row(ws, "REVENUE", "All revenue lines — formula-driven from ASMP & SIZE")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, formula, fmt=FMT_CR, bold=False, fill=None):
        ws.cell(row=row, column=COL_YR0+j).value = formula
        ws.cell(row=row, column=COL_YR0+j).number_format = fmt
        ws.cell(row=row, column=COL_YR0+j).font = _font(bold=bold)
        if fill: ws.cell(row=row, column=COL_YR0+j).fill = _fill(fill)

    # ── Escalated rates ───────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "ESCALATED RATES (MEMO)"); r += 1
    REV_R['rack_mrc_esc'] = r
    _lbl(ws, r, "Rack MRC (escalated)", "Cr/rack/mo")
    for j in range(N):
        f(r, j, f"={_asmp('rack_mrc')}*(1+{_asmp('rack_mrc_esc')})^({cl(j)}4-1)", FMT_CR)
    r += 1

    REV_R['otc_esc'] = r
    _lbl(ws, r, "OTC fee per rack (escalated)", "Cr/rack")
    for j in range(N):
        f(r, j, f"={_asmp('otc_fee')}*(1+{_asmp('otc_esc')})^({cl(j)}4-1)", FMT_CR)
    r += 1

    REV_R['tenant_esc'] = r
    _lbl(ws, r, "Tenant power tariff (escalated)", "Rs/kWh")
    for j in range(N):
        f(r, j, f"={_asmp('tenant_tariff')}*(1+{_asmp('pwr_esc')})^({cl(j)}4-1)", FMT_CR)
    r += 1

    REV_R['util_esc'] = r
    _lbl(ws, r, "Grid tariff (escalated, cost to DC)", "Rs/kWh")
    for j in range(N):
        f(r, j, f"={_asmp('util_tariff')}*(1+{_asmp('pwr_esc')})^({cl(j)}4-1)", FMT_CR)
    r += 1

    # ── Revenue lines ─────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "REVENUE BUILD"); r += 1
    occ = SIZE_R['occupied']
    REV_R['colo'] = r
    _lbl(ws, r, "Recurring colo revenue", "Cr")
    for j in range(N):
        f(r, j, f"=SIZE!{cl(j)}{occ}*REV!{cl(j)}{REV_R['rack_mrc_esc']}*12", FMT_CR)
    r += 1

    REV_R['otc'] = r
    _lbl(ws, r, "OTC setup revenue", "Cr")
    for j in range(N):
        f(r, j, f"=SIZE!{cl(j)}{SIZE_R['new_racks']}*REV!{cl(j)}{REV_R['otc_esc']}", FMT_CR)
    r += 1

    REV_R['power'] = r
    _lbl(ws, r, "Power revenue", "Cr")
    for j in range(N):
        f(r, j, f"=SIZE!{cl(j)}{SIZE_R['fac_load']}*REV!{cl(j)}{REV_R['tenant_esc']}*8760/10000000", FMT_CR)
    r += 1

    r += 1  # blank
    REV_R['gross'] = r
    _lbl(ws, r, "Gross revenue", "Cr", bold=True)
    for j in range(N):
        f(r, j, f"=REV!{cl(j)}{REV_R['colo']}+REV!{cl(j)}{REV_R['otc']}+REV!{cl(j)}{REV_R['power']}", FMT_CR, bold=True, fill=LTGREY)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, bold=True, fill=LTGREY)
    r += 1

    REV_R['dot'] = r
    _lbl(ws, r, "DoT revenue share (deduction)", "Cr")
    for j in range(N):
        f(r, j, f"=-REV!{cl(j)}{REV_R['gross']}*{_asmp('dot_share')}", FMT_CR)
    r += 1

    REV_R['net'] = r
    _lbl(ws, r, "Net revenue", "Cr", bold=True)
    for j in range(N):
        f(r, j, f"=REV!{cl(j)}{REV_R['gross']}+REV!{cl(j)}{REV_R['dot']}", FMT_CR, bold=True, fill=LTGREY)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, bold=True, fill=LTGREY)
    r += 2

    # ── Power cost memo ───────────────────────────────────────────────────
    _subhdr(ws, r, "POWER COST & MARGIN (MEMO)"); r += 1
    REV_R['pwr_cost'] = r
    _lbl(ws, r, "Power cost (grid tariff × load)", "Cr")
    for j in range(N):
        f(r, j, f"=SIZE!{cl(j)}{SIZE_R['fac_load']}*REV!{cl(j)}{REV_R['util_esc']}*8760/10000000", FMT_CR)
    r += 1

    REV_R['pwr_margin'] = r
    _lbl(ws, r, "Power margin", "Cr")
    for j in range(N):
        f(r, j, f"=REV!{cl(j)}{REV_R['power']}-REV!{cl(j)}{REV_R['pwr_cost']}", FMT_CR)
    r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# CAPEX
# ═══════════════════════════════════════════════════════════════════════════════
def write_capex(wb):
    ws = wb.create_sheet("CAPEX")
    _col_widths(ws); _title_row(ws, "CAPITAL EXPENDITURE", "Phased deployment — formula-driven from ASMP")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, val_or_fml, fmt=FMT_CR, bold=False, fill=None):
        c = ws.cell(row=row, column=COL_YR0+j)
        c.value = val_or_fml
        c.number_format = fmt
        c.font = _font(bold=bold)
        if fill: c.fill = _fill(fill)

    dep = f"ASMP!{{}}{AR['deploy_arr']}"  # template for deploy array ref

    # Helper: per-rack × racks_this_phase formula
    def pr_fml(asmp_name, j):
        return f"={_asmp(asmp_name)}*ASMP!{cl(j)}${AR['deploy_arr']}"

    r += 1; _hdr(ws, r, "CAPEX COMPONENTS"); r += 1

    CAP_R['civil'] = r
    _lbl(ws, r, "Civil & structural (shell — full build, Ph1)", "Cr")
    # Shell is built once for full capacity in year 1 (civil/rack × total racks),
    # not phased — matches the engine's shell/fit-out split. Fit-out below stays phased.
    for j in range(N):
        if j == 0:
            f(r, j, f"={_asmp('civil_pr')}*SUM(ASMP!{cl(0)}${AR['deploy_arr']}:{cl(N-1)}${AR['deploy_arr']})")
        else:
            f(r, j, 0)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CAP_R['elec'] = r
    _lbl(ws, r, "Electrical systems", "Cr")
    for j in range(N): f(r, j, pr_fml('elec_pr', j))
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CAP_R['mech'] = r
    _lbl(ws, r, "Mechanical systems", "Cr")
    for j in range(N): f(r, j, pr_fml('mech_pr', j))
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CAP_R['it'] = r
    _lbl(ws, r, "IT hardware", "Cr")
    for j in range(N): f(r, j, pr_fml('it_pr', j))
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CAP_R['net'] = r
    _lbl(ws, r, "Network infrastructure", "Cr")
    for j in range(N): f(r, j, pr_fml('net_pr', j))
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CAP_R['software'] = r
    _lbl(ws, r, "Software & licensing (Ph1 only)", "Cr")
    for j in range(N):
        f(r, j, f"={_asmp('software_c')}" if j == 0 else 0)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CAP_R['site'] = r
    _lbl(ws, r, "Land & site prep (Ph1 only)", "Cr")
    for j in range(N):
        f(r, j, f"={_asmp('land_c')}+{_asmp('siteprep_c')}+{_asmp('misc_infra_c')}" if j == 0 else 0)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CAP_R['preop'] = r
    _lbl(ws, r, "Pre-operational expenses", "Cr")
    hard_rows = [CAP_R['civil'], CAP_R['elec'], CAP_R['mech'], CAP_R['it'], CAP_R['net']]
    for j in range(N):
        hard = "+".join(f"CAPEX!{cl(j)}{rr}" for rr in hard_rows)
        f(r, j, f"={_asmp('preop_pct')}*({hard})")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    r += 1
    CAP_R['total'] = r
    _lbl(ws, r, "Total CapEx", "Cr", bold=True)
    all_rows = [CAP_R['civil'], CAP_R['elec'], CAP_R['mech'], CAP_R['it'],
                CAP_R['net'], CAP_R['software'], CAP_R['site'], CAP_R['preop']]
    for j in range(N):
        terms = "+".join(f"CAPEX!{cl(j)}{rr}" for rr in all_rows)
        f(r, j, f"={terms}", bold=True, fill=LTGREY)
        ws.cell(row=r, column=COL_YR0+j).number_format = FMT_CR
        ws.cell(row=r, column=COL_YR0+j).border = _bdr(True, True)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, bold=True, fill=LTGREY, bdr=True)
    _lbl(ws, r, "Total CapEx", "Cr", bold=True, fill=LTGREY); r += 1

    CAP_R['cumul'] = r
    _lbl(ws, r, "Cumulative CapEx", "Cr", bold=True)
    for j in range(N):
        if j == 0:
            f(r, j, f"=CAPEX!{cl(0)}{CAP_R['total']}", FMT_CR, True, LTGREY)
        else:
            f(r, j, f"=CAPEX!{cl(j-1)}{CAP_R['cumul']}+CAPEX!{cl(j)}{CAP_R['total']}", FMT_CR, True, LTGREY)
    r += 2

    # ── Cumulative by component (for DEPR & AMC) ──────────────────────────
    _subhdr(ws, r, "CUMULATIVE BY COMPONENT (for DEPR & AMC)"); r += 1

    for key, label in [('civil_c','Civil'),('elec_c','Electrical'),
                       ('mech_c','Mechanical'),('net_c','Network'),('soft_c','Software')]:
        src = {'civil_c': CAP_R['civil'], 'elec_c': CAP_R['elec'], 'mech_c': CAP_R['mech'],
               'net_c': CAP_R['net'], 'soft_c': CAP_R['software']}[key]
        CAP_R[key] = r
        _lbl(ws, r, f"Cumulative {label}", "Cr")
        for j in range(N):
            if j == 0:
                f(r, j, f"=CAPEX!{cl(0)}{src}")
            else:
                f(r, j, f"=CAPEX!{cl(j-1)}{CAP_R[key]}+CAPEX!{cl(j)}{src}")
        r += 1

    r += 1
    # ── Maintenance CapEx — M&E base (electrical + mechanical), 3-yr OEM warranty ──
    # Base is the maintainable M&E plant only (matches engine); each phase becomes
    # eligible maint_warranty years after commissioning. Period is 1-based (row 4),
    # so the threshold is MAX(deploy_idx, construction_years) + warranty + 1.
    CAP_R['maint'] = r
    _lbl(ws, r, "Maintenance CapEx (M&E base, post-OEM warranty)", "Cr")
    for j in range(N):
        eligible_terms = "+".join(
            f"IF({cl(j)}4>=MAX({DEPLOY_XIDX[p]-1},{_asmp('construction_years')})+{_asmp('maint_warranty')}+1,"
            f"CAPEX!${DEPLOY_XCOL[p]}${CAP_R['elec']}+CAPEX!${DEPLOY_XCOL[p]}${CAP_R['mech']},0)"
            for p in range(3)
        )
        f(r, j, f"={_asmp('maint_cx')}*({eligible_terms})")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 2

    # ── Sources & Uses ────────────────────────────────────────────────────
    _subhdr(ws, r, "SOURCES & USES"); r += 1
    CAP_R['eq_fund'] = r
    _lbl(ws, r, "Equity funded", "Cr")
    for j in range(N):
        f(r, j, f"=CAPEX!{cl(j)}{CAP_R['total']}*{_asmp('eq_pct')}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CAP_R['debt_fund'] = r
    _lbl(ws, r, "Debt funded", "Cr")
    for j in range(N):
        f(r, j, f"=CAPEX!{cl(j)}{CAP_R['total']}*{_asmp('debt_pct')}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# OPEX
# ═══════════════════════════════════════════════════════════════════════════════
def write_opex(wb):
    ws = wb.create_sheet("OPEX")
    _col_widths(ws); _title_row(ws, "OPEX & EBITDA", "Formula-driven from ASMP, SIZE, REV, CAPEX")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, formula, fmt=FMT_CR, bold=False, fill=None):
        c = ws.cell(row=row, column=COL_YR0+j)
        c.value = formula; c.number_format = fmt; c.font = _font(bold=bold)
        if fill: c.fill = _fill(fill)

    occ = SIZE_R['occupied']; net = REV_R['net']

    r += 1; _hdr(ws, r, "OPEX LINES"); r += 1

    OPX_R['power'] = r
    _lbl(ws, r, "Power cost (grid tariff × load)", "Cr")
    for j in range(N):
        f(r, j, f"=SIZE!{cl(j)}{SIZE_R['fac_load']}*REV!{cl(j)}{REV_R['util_esc']}*8760/10000000")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    OPX_R['manpower'] = r
    _lbl(ws, r, "Manpower", "Cr")
    for j in range(N):
        f(r, j, f"=ROUNDUP(SIZE!{cl(j)}{occ}/100*{_asmp('fte_per100')},0)"
                f"*{_asmp('avg_ctc')}/100*(1+{_asmp('mp_esc')})^({cl(j)}4-1)")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    # Construction-year gate: these costs are zero until operations begin
    # (period > construction_years), matching the engine.
    cy_gate = lambda j: f"{cl(j)}4>{_asmp('construction_years')}"

    OPX_R['housekeeping'] = r
    _lbl(ws, r, "Housekeeping", "Cr")
    for j in range(N):
        f(r, j, f"=IF({cy_gate(j)},{_asmp('facility_sqft')}*{_asmp('hk_rate')}*(1+{_asmp('hk_esc')})^({cl(j)}4-1)/10000000,0)")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    OPX_R['amc'] = r
    _lbl(ws, r, "Maintenance (AMC, asset-based)", "Cr")
    for j in range(N):
        c = cl(j)
        f(r, j, f"=IF({cy_gate(j)},{_asmp('amc_civil')}*CAPEX!{c}{CAP_R['civil_c']}"
                f"+{_asmp('amc_elec')}*CAPEX!{c}{CAP_R['elec_c']}"
                f"+{_asmp('amc_mech')}*CAPEX!{c}{CAP_R['mech_c']}"
                f"+{_asmp('amc_net')}*CAPEX!{c}{CAP_R['net_c']}"
                f"+{_asmp('amc_soft')}*CAPEX!{c}{CAP_R['soft_c']},0)")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    OPX_R['network'] = r
    _lbl(ws, r, "Network & connectivity", "Cr")
    for j in range(N):
        f(r, j, f"={_asmp('net_pct')}*REV!{cl(j)}{net}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    OPX_R['security'] = r
    _lbl(ws, r, "Security (physical)", "Cr")
    for j in range(N):
        f(r, j, f"={_asmp('sec_pct')}*REV!{cl(j)}{net}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    OPX_R['insurance'] = r
    _lbl(ws, r, "Insurance", "Cr")
    for j in range(N):
        f(r, j, f"=IF({cy_gate(j)},{_asmp('ins_pct')}*CAPEX!{cl(j)}{CAP_R['cumul']},0)")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    OPX_R['ptax'] = r
    _lbl(ws, r, "Property tax (land + civil basis)", "Cr")
    for j in range(N):
        base = f"({_asmp('land_c')}+SUM(CAPEX!$C${CAP_R['civil']}:{cl(j)}${CAP_R['civil']}))"
        f(r, j, f"=IF({cy_gate(j)},{_asmp('ptax_pct')}*{base},0)")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    OPX_R['marketing'] = r
    _lbl(ws, r, "Marketing (declining % of net revenue)", "Cr")
    for j in range(N):
        pct = (f"MAX({_asmp('mkt_pct_end')},"
               f"{_asmp('mkt_pct_start')}"
               f"-({_asmp('mkt_pct_start')}-{_asmp('mkt_pct_end')})"
               f"*({cl(j)}4-1)/({_asmp('proj_years')}-1))")
        f(r, j, f"={pct}*REV!{cl(j)}{REV_R['net']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    OPX_R['gna'] = r
    _lbl(ws, r, "G&A", "Cr")
    for j in range(N):
        f(r, j, f"={_asmp('gna_pct')}*REV!{cl(j)}{net}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    r += 1
    opx_rows = [OPX_R[k] for k in ('power','manpower','housekeeping','amc','network',
                                    'security','insurance','ptax','marketing','gna')]
    OPX_R['total'] = r
    _lbl(ws, r, "Total OpEx", "Cr", bold=True, fill=LTGREY)
    for j in range(N):
        terms = "+".join(f"OPEX!{cl(j)}{rr}" for rr in opx_rows)
        f(r, j, f"={terms}", FMT_CR, True, LTGREY)
        ws.cell(row=r, column=COL_YR0+j).border = _bdr(True, True)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True, LTGREY, bdr=True)
    r += 2

    OPX_R['ebitda'] = r
    _lbl(ws, r, "EBITDA", "Cr", bold=True, fill=LTGREY)
    for j in range(N):
        f(r, j, f"=REV!{cl(j)}{REV_R['net']}-OPEX!{cl(j)}{OPX_R['total']}", FMT_CR, True, LTGREY)
        ws.cell(row=r, column=COL_YR0+j).border = _bdr(True, True)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True, LTGREY, bdr=True)
    r += 1

    OPX_R['ebitda_m'] = r
    _lbl(ws, r, "EBITDA margin", "%")
    for j in range(N):
        f(r, j, f"=IF(REV!{cl(j)}{REV_R['net']}>0,OPEX!{cl(j)}{OPX_R['ebitda']}/REV!{cl(j)}{REV_R['net']},0)", FMT_P1)
    r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# DEPR — Book depreciation (SLM)
# ═══════════════════════════════════════════════════════════════════════════════
def write_depr(wb):
    ws = wb.create_sheet("DEPR")
    _col_widths(ws); _title_row(ws, "DEPRECIATION", "SLM — Companies Act useful lives from ASMP")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, formula, fmt=FMT_CR, bold=False, fill=None):
        c = ws.cell(row=row, column=COL_YR0+j)
        c.value = formula; c.number_format = fmt; c.font = _font(bold=bold)
        if fill: c.fill = _fill(fill)

    def dep_fml(j, capex_row, life_key):
        """SLM dep for year j = sum over phases of IF(period >= commissioning, capex_phase/life, 0).
        Depreciation begins at commissioning = MAX(deploy_idx_0based, construction_years)+1
        (1-based period), so assets deployed in the construction year aren't depreciated
        until operations begin — matches the engine's dep_start = max(purchase_yr, constr_yrs)."""
        parts = []
        for ph_col, ph_idx in zip(DEPLOY_XCOL, DEPLOY_XIDX):
            parts.append(
                f"IF({cl(j)}4>=MAX({ph_idx-1},{_asmp('construction_years')})+1,"
                f"CAPEX!${ph_col}${capex_row}/{_asmp(life_key)},0)"
            )
        return "=" + "+".join(parts)

    r += 1; _hdr(ws, r, "BOOK DEPRECIATION (SLM)"); r += 1

    DEP_R['civil'] = r
    _lbl(ws, r, "Civil & structural", "Cr")
    for j in range(N): f(r, j, dep_fml(j, CAP_R['civil'], 'life_civil'))
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    DEP_R['elec'] = r
    _lbl(ws, r, "Electrical systems", "Cr")
    for j in range(N): f(r, j, dep_fml(j, CAP_R['elec'], 'life_elec'))
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    DEP_R['mech'] = r
    _lbl(ws, r, "Mechanical systems", "Cr")
    for j in range(N): f(r, j, dep_fml(j, CAP_R['mech'], 'life_mech'))
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    DEP_R['it'] = r
    _lbl(ws, r, "IT hardware", "Cr")
    for j in range(N): f(r, j, dep_fml(j, CAP_R['it'], 'life_it'))
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    DEP_R['net'] = r
    _lbl(ws, r, "Network infrastructure", "Cr")
    for j in range(N): f(r, j, dep_fml(j, CAP_R['net'], 'life_net'))
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    DEP_R['soft'] = r
    _lbl(ws, r, "Software & licensing", "Cr")
    for j in range(N):
        # Software deployed in Phase 1 (col C); deferred to commissioning like other assets
        f(r, j, f"=IF({cl(j)}4>={_asmp('construction_years')}+1,CAPEX!$C${CAP_R['software']}/{_asmp('life_soft')},0)")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    r += 1
    dep_rows = [DEP_R[k] for k in ('civil','elec','mech','it','net','soft')]
    DEP_R['total'] = r
    _lbl(ws, r, "Total depreciation", "Cr", bold=True, fill=LTGREY)
    for j in range(N):
        terms = "+".join(f"DEPR!{cl(j)}{rr}" for rr in dep_rows)
        f(r, j, f"={terms}", FMT_CR, True, LTGREY)
        ws.cell(row=r, column=COL_YR0+j).border = _bdr(True, True)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True, LTGREY, bdr=True)
    r += 1

    DEP_R['accum'] = r
    _lbl(ws, r, "Accumulated depreciation", "Cr", bold=True)
    for j in range(N):
        if j == 0:
            f(r, j, f"=DEPR!{cl(0)}{DEP_R['total']}", FMT_CR, True, LTGREY)
        else:
            f(r, j, f"=DEPR!{cl(j-1)}{DEP_R['accum']}+DEPR!{cl(j)}{DEP_R['total']}", FMT_CR, True, LTGREY)
    r += 1

    DEP_R['nbv'] = r
    _lbl(ws, r, "Net book value (NBV)", "Cr", bold=True)
    for j in range(N):
        f(r, j, f"=CAPEX!{cl(j)}{CAP_R['cumul']}-DEPR!{cl(j)}{DEP_R['accum']}", FMT_CR, True, LTGREY)
    _w(ws, r, COL_YR0+N, f"=DEPR!L{r}", FMT_CR, True, LTGREY)  # terminal NBV
    r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# DEBT — Loan schedule
# ═══════════════════════════════════════════════════════════════════════════════
def write_debt(wb):
    ws = wb.create_sheet("DEBT")
    _col_widths(ws); _title_row(ws, "DEBT SCHEDULE", "3 tranches — equal principal repayment")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, formula, fmt=FMT_CR, bold=False, fill=None):
        c = ws.cell(row=row, column=COL_YR0+j)
        c.value = formula; c.number_format = fmt; c.font = _font(bold=bold)
        if fill: c.fill = _fill(fill)

    tranche_cfg = [
        (1, 0, 'C', 1),   # tranche, draw_py_idx, draw_col, draw_excel_idx
        (2, 3, 'F', 4),
        (3, 6, 'I', 7),
    ]

    t_loan_rows = []  # loan amount scalar row per tranche

    for tranche_no, draw_py, draw_xcol, draw_xidx in tranche_cfg:
        r += 1
        _subhdr(ws, r, f"TRANCHE {tranche_no}  (drawn {YEARS[draw_py]})", 13); r += 1

        loan_r = r
        t_loan_rows.append(loan_r)
        _lbl(ws, r, f"Loan amount (Tranche {tranche_no})", "Cr", bold=True)
        # Loan amount only in deployment column
        for j in range(N):
            if cl(j) == draw_xcol:
                f(r, j, f"=CAPEX!${draw_xcol}${CAP_R['total']}*{_asmp('debt_pct')}", FMT_CR, True, BLUE_IN)
            else:
                f(r, j, 0, FMT_CR)
        r += 1

        # Repayment condition: year_index >= draw_xidx + moratorium + 1 (formula-driven via ASMP)
        rep_cond_tpl = f"{{c}}4>={draw_xidx}+{_asmp('morat')}+1"
        # Amortize over the post-moratorium repayment period (tenure − moratorium)
        # so the loan repays at the correct rate to mature on schedule (#5).
        annual_princ_fml = f"DEBT!${draw_xcol}${loan_r}/({_asmp('tenure')}-{_asmp('morat')})"

        open_r = r; _lbl(ws, r, "  Opening balance", "Cr"); r += 1
        draw_r = r; _lbl(ws, r, "  Drawdown", "Cr"); r += 1
        int_r  = r; _lbl(ws, r, "  Interest expense", "Cr"); r += 1
        prin_r = r; _lbl(ws, r, "  Principal repayment", "Cr"); r += 1
        clos_r = r; _lbl(ws, r, "  Closing balance", "Cr", bold=True); r += 1

        for j in range(N):
            c = cl(j)
            # Opening
            if j == 0:
                f(open_r, j, 0)
            else:
                f(open_r, j, f"=DEBT!{cl(j-1)}{clos_r}")
            # Drawdown
            f(draw_r, j, f"=DEBT!{c}{loan_r}")
            # Interest = rate × (opening + drawdown); half-year in the draw
            # year since the tranche is drawn progressively, not day one (#6).
            if j == draw_py:
                f(int_r, j, f"={_asmp('int_rate')}*(DEBT!{c}{open_r}+DEBT!{c}{draw_r})*0.5")
            else:
                f(int_r, j, f"={_asmp('int_rate')}*(DEBT!{c}{open_r}+DEBT!{c}{draw_r})")
            # Principal
            rep_cond = rep_cond_tpl.format(c=c)
            f(prin_r, j, f"=IF({rep_cond},MIN({annual_princ_fml},DEBT!{c}{open_r}+DEBT!{c}{draw_r}),0)")
            # Closing
            f(clos_r, j, f"=DEBT!{c}{open_r}+DEBT!{c}{draw_r}-DEBT!{c}{prin_r}", fill=LTGREY)

        DBT_R[f't{tranche_no}_int']  = int_r
        DBT_R[f't{tranche_no}_prin'] = prin_r
        DBT_R[f't{tranche_no}_clos'] = clos_r

    # ── Consolidated ──────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "CONSOLIDATED DEBT ACCOUNT"); r += 1

    DBT_R['drawdown'] = r
    _lbl(ws, r, "Total drawdown", "Cr")
    for j in range(N):
        # Sum of all loan_amount rows in that year column
        terms = "+".join(f"DEBT!{cl(j)}{lr}" for lr in t_loan_rows)
        f(r, j, f"={terms}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    DBT_R['equity'] = r
    _lbl(ws, r, "Equity injection", "Cr")
    for j in range(N):
        f(r, j, f"=CAPEX!{cl(j)}{CAP_R['eq_fund']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    DBT_R['interest'] = r
    _lbl(ws, r, "Total interest expense", "Cr")
    for j in range(N):
        terms = "+".join(f"DEBT!{cl(j)}{DBT_R[f't{i}_int']}" for i in range(1,4))
        f(r, j, f"={terms}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    DBT_R['principal'] = r
    _lbl(ws, r, "Total principal repayment", "Cr")
    for j in range(N):
        terms = "+".join(f"DEBT!{cl(j)}{DBT_R[f't{i}_prin']}" for i in range(1,4))
        f(r, j, f"={terms}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    DBT_R['closing'] = r
    _lbl(ws, r, "Closing debt balance", "Cr", bold=True)
    for j in range(N):
        terms = "+".join(f"DEBT!{cl(j)}{DBT_R[f't{i}_clos']}" for i in range(1,4))
        f(r, j, f"={terms}", bold=True, fill=LTGREY)
        ws.cell(row=r, column=COL_YR0+j).border = _bdr(True, True)
    _w(ws, r, COL_YR0+N, f"=DEBT!L{r}", FMT_CR, True, LTGREY, bdr=True)
    r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# TAX — Simplified (EBT, Cumulative EBT, Tax)
# ═══════════════════════════════════════════════════════════════════════════════
def write_tax(wb):
    ws = wb.create_sheet("TAX")
    _col_widths(ws)
    _title_row(ws, "TAXATION",
               "Tax accrues when cumulative taxable profit turns positive (loss carry-forward)")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, formula, fmt=FMT_CR, bold=False, fill=None):
        c = ws.cell(row=row, column=COL_YR0+j)
        c.value = formula; c.number_format = fmt; c.font = _font(bold=bold)
        if fill: c.fill = _fill(fill)

    r += 1; _hdr(ws, r, "TAX SCHEDULE"); r += 1

    TAX_R['ebt'] = r
    _lbl(ws, r, "EBT", "Cr")
    for j in range(N):
        f(r, j, f"=PNL!{cl(j)}{PNL_R['ebt']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    TAX_R['cumul_ebt'] = r
    _lbl(ws, r, "Cumulative taxable profit", "Cr")
    for j in range(N):
        if j == 0:
            f(r, j, f"=TAX!{cl(0)}{TAX_R['ebt']}")
        else:
            f(r, j, f"=TAX!{cl(j-1)}{TAX_R['cumul_ebt']}+TAX!{cl(j)}{TAX_R['ebt']}")
    r += 1

    TAX_R['tax'] = r
    _lbl(ws, r, "Income tax payable", "Cr", bold=True)
    for j in range(N):
        cebt = TAX_R['cumul_ebt']
        if j == 0:
            f(r, j, f"=MAX(TAX!{cl(0)}{cebt},0)*{_asmp('tax_rate')}", bold=True, fill=LTGREY)
        else:
            f(r, j,
              f"=(MAX(TAX!{cl(j)}{cebt},0)-MAX(TAX!{cl(j-1)}{cebt},0))*{_asmp('tax_rate')}",
              FMT_CR, True, LTGREY)
        ws.cell(row=r, column=COL_YR0+j).border = _bdr(True, True)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True, LTGREY, bdr=True)
    r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# WC — Working capital
# ═══════════════════════════════════════════════════════════════════════════════
def write_wc(wb):
    ws = wb.create_sheet("WC")
    _col_widths(ws); _title_row(ws, "WORKING CAPITAL", "Net working capital & annual movement")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, formula, fmt=FMT_CR, bold=False, fill=None):
        c = ws.cell(row=row, column=COL_YR0+j)
        c.value = formula; c.number_format = fmt; c.font = _font(bold=bold)
        if fill: c.fill = _fill(fill)

    r += 1; _hdr(ws, r, "WORKING CAPITAL (DSO / DPO method)"); r += 1

    WC_R['receivables'] = r
    _lbl(ws, r, "Accounts receivable (Net rev × DSO/365)", "Cr")
    for j in range(N):
        f(r, j, f"=REV!{cl(j)}{REV_R['net']}*{_asmp('dso')}/365")
    r += 1

    WC_R['payables'] = r
    _lbl(ws, r, "Accounts payable (Cash opex × DPO/365)", "Cr")
    for j in range(N):
        f(r, j, f"=OPEX!{cl(j)}{OPX_R['total']}*{_asmp('dpo')}/365")
    r += 1

    WC_R['nwc'] = r
    _lbl(ws, r, "Net working capital (AR − AP)", "Cr", bold=True)
    for j in range(N):
        f(r, j, f"=WC!{cl(j)}{WC_R['receivables']}-WC!{cl(j)}{WC_R['payables']}", FMT_CR, True, LTGREY)
    r += 1

    WC_R['delta'] = r
    _lbl(ws, r, "Change in working capital (ΔWC)", "Cr", bold=True)
    for j in range(N):
        if j == 0:
            f(r, j, f"=WC!{cl(0)}{WC_R['nwc']}", FMT_CR, True, LTGREY)
        else:
            f(r, j, f"=WC!{cl(j)}{WC_R['nwc']}-WC!{cl(j-1)}{WC_R['nwc']}", FMT_CR, True, LTGREY)
    r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# PNL — Income statement
# ═══════════════════════════════════════════════════════════════════════════════
def write_pnl(wb):
    ws = wb.create_sheet("PNL")
    _col_widths(ws); _title_row(ws, "PROFIT & LOSS", "Full income statement — all formula-driven")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, formula, fmt=FMT_CR, bold=False, fill=None, pct=False):
        c = ws.cell(row=row, column=COL_YR0+j)
        c.value = formula; c.font = _font(bold=bold)
        c.number_format = FMT_P1 if pct else fmt
        if fill: c.fill = _fill(fill)

    r += 1; _hdr(ws, r, "INCOME STATEMENT"); r += 1

    PNL_R['net_rev'] = r
    _lbl(ws, r, "Net revenue", "Cr", bold=True)
    for j in range(N):
        f(r, j, f"=REV!{cl(j)}{REV_R['net']}", bold=True)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True); r += 1

    PNL_R['total_opx'] = r
    _lbl(ws, r, "Total OpEx", "Cr")
    for j in range(N):
        f(r, j, f"=-OPEX!{cl(j)}{OPX_R['total']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    PNL_R['ebitda'] = r
    _lbl(ws, r, "EBITDA", "Cr", bold=True, fill=LTGREY)
    for j in range(N):
        f(r, j, f"=PNL!{cl(j)}{PNL_R['net_rev']}+PNL!{cl(j)}{PNL_R['total_opx']}", bold=True, fill=LTGREY)
        ws.cell(row=r, column=COL_YR0+j).number_format = FMT_CR
        ws.cell(row=r, column=COL_YR0+j).border = _bdr(True, True)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True, LTGREY, bdr=True); r += 1

    PNL_R['ebitda_m'] = r
    _lbl(ws, r, "EBITDA margin", "%", italic=True)
    for j in range(N):
        f(r, j, f"=IF(PNL!{cl(j)}{PNL_R['net_rev']}>0,"
                f"PNL!{cl(j)}{PNL_R['ebitda']}/PNL!{cl(j)}{PNL_R['net_rev']},0)", pct=True)
    r += 1

    PNL_R['dep'] = r
    _lbl(ws, r, "Depreciation (SLM)", "Cr")
    for j in range(N):
        f(r, j, f"=-DEPR!{cl(j)}{DEP_R['total']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    PNL_R['ebit'] = r
    _lbl(ws, r, "EBIT", "Cr", bold=True, fill=LTGREY)
    for j in range(N):
        f(r, j, f"=PNL!{cl(j)}{PNL_R['ebitda']}+PNL!{cl(j)}{PNL_R['dep']}", bold=True, fill=LTGREY)
        ws.cell(row=r, column=COL_YR0+j).number_format = FMT_CR
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True, LTGREY); r += 1

    PNL_R['interest'] = r
    _lbl(ws, r, "Interest expense", "Cr")
    for j in range(N):
        f(r, j, f"=-DEBT!{cl(j)}{DBT_R['interest']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    PNL_R['ebt'] = r
    _lbl(ws, r, "EBT / PBT", "Cr", bold=True, fill=LTGREY)
    for j in range(N):
        f(r, j, f"=PNL!{cl(j)}{PNL_R['ebit']}+PNL!{cl(j)}{PNL_R['interest']}", bold=True, fill=LTGREY)
        ws.cell(row=r, column=COL_YR0+j).number_format = FMT_CR
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True, LTGREY); r += 1

    PNL_R['tax'] = r
    _lbl(ws, r, "Income tax", "Cr")
    for j in range(N):
        f(r, j, f"=-TAX!{cl(j)}{TAX_R['tax']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    PNL_R['pat'] = r
    _lbl(ws, r, "PAT (Profit after tax)", "Cr", bold=True, fill=LTGREY)
    for j in range(N):
        f(r, j, f"=PNL!{cl(j)}{PNL_R['ebt']}+PNL!{cl(j)}{PNL_R['tax']}", bold=True, fill=LTGREY)
        ws.cell(row=r, column=COL_YR0+j).number_format = FMT_CR
        ws.cell(row=r, column=COL_YR0+j).border = _bdr(True, True)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True, LTGREY, bdr=True); r += 1

    PNL_R['pat_m'] = r
    _lbl(ws, r, "PAT margin", "%", italic=True)
    for j in range(N):
        f(r, j, f"=IF(PNL!{cl(j)}{PNL_R['net_rev']}>0,"
                f"PNL!{cl(j)}{PNL_R['pat']}/PNL!{cl(j)}{PNL_R['net_rev']},0)", pct=True)
    r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# CFS — Cash flow statement
# ═══════════════════════════════════════════════════════════════════════════════
def write_cfs(wb):
    ws = wb.create_sheet("CFS")
    _col_widths(ws); _title_row(ws, "CASH FLOW STATEMENT", "FCFF, FCFE, CFADS & DSCR")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, formula, fmt=FMT_CR, bold=False, fill=None):
        c = ws.cell(row=row, column=COL_YR0+j)
        c.value = formula; c.number_format = fmt; c.font = _font(bold=bold)
        if fill: c.fill = _fill(fill)

    r += 1; _hdr(ws, r, "PROJECT FREE CASH FLOW (FCFF)"); r += 1

    CFS_R['nopat'] = r
    _lbl(ws, r, "NOPAT  (EBIT × (1 - tax rate))", "Cr")
    for j in range(N):
        f(r, j, f"=PNL!{cl(j)}{PNL_R['ebit']}*(1-{_asmp('tax_rate')})")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CFS_R['dep_add'] = r
    _lbl(ws, r, "Add: depreciation (non-cash)", "Cr")
    for j in range(N):
        f(r, j, f"=DEPR!{cl(j)}{DEP_R['total']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CFS_R['delta_wc'] = r
    _lbl(ws, r, "Less: increase in working capital", "Cr")
    for j in range(N):
        f(r, j, f"=-WC!{cl(j)}{WC_R['delta']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CFS_R['capex'] = r
    _lbl(ws, r, "Less: CapEx (growth)", "Cr")
    for j in range(N):
        f(r, j, f"=-CAPEX!{cl(j)}{CAP_R['total']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CFS_R['maint_cx'] = r
    _lbl(ws, r, "Less: Maintenance CapEx", "Cr")
    for j in range(N):
        f(r, j, f"=-CAPEX!{cl(j)}{CAP_R['maint']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    r += 1
    CFS_R['fcff'] = r
    _lbl(ws, r, "FCFF", "Cr", bold=True, fill=LTGREY)
    for j in range(N):
        terms = "+".join(f"CFS!{cl(j)}{CFS_R[k]}"
                         for k in ('nopat','dep_add','delta_wc','capex','maint_cx'))
        f(r, j, f"={terms}", FMT_CR, True, LTGREY)
        ws.cell(row=r, column=COL_YR0+j).border = _bdr(True, True)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True, LTGREY, bdr=True); r += 1

    CFS_R['cumul_fcff'] = r
    _lbl(ws, r, "Cumulative FCFF", "Cr")
    for j in range(N):
        if j == 0:
            f(r, j, f"=CFS!{cl(0)}{CFS_R['fcff']}")
        else:
            f(r, j, f"=CFS!{cl(j-1)}{CFS_R['cumul_fcff']}+CFS!{cl(j)}{CFS_R['fcff']}")
    r += 2

    # ── Equity cashflow ────────────────────────────────────────────────────
    _hdr(ws, r, "EQUITY FREE CASH FLOW (FCFE)"); r += 1

    CFS_R['drawdown'] = r
    _lbl(ws, r, "Add: debt drawdown", "Cr")
    for j in range(N):
        f(r, j, f"=DEBT!{cl(j)}{DBT_R['drawdown']}")
    r += 1

    CFS_R['eq_inj'] = r
    _lbl(ws, r, "Less: equity injection", "Cr")
    for j in range(N):
        f(r, j, f"=-DEBT!{cl(j)}{DBT_R['equity']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CFS_R['principal'] = r
    _lbl(ws, r, "Less: principal repayment", "Cr")
    for j in range(N):
        f(r, j, f"=-DEBT!{cl(j)}{DBT_R['principal']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CFS_R['int_paid'] = r
    _lbl(ws, r, "Less: interest paid (post-tax shield)", "Cr")
    for j in range(N):
        f(r, j, f"=-DEBT!{cl(j)}{DBT_R['interest']}*(1-{_asmp('tax_rate')})")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    r += 1
    CFS_R['fcfe'] = r
    _lbl(ws, r, "FCFE", "Cr", bold=True, fill=LTGREY)
    for j in range(N):
        terms = "+".join(f"CFS!{cl(j)}{CFS_R[k]}"
                         for k in ('fcff','drawdown','eq_inj','principal','int_paid'))
        f(r, j, f"={terms}", FMT_CR, True, LTGREY)
        ws.cell(row=r, column=COL_YR0+j).border = _bdr(True, True)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True, LTGREY, bdr=True); r += 2

    # ── CFADS & DSCR ──────────────────────────────────────────────────────
    _hdr(ws, r, "DEBT SERVICE COVERAGE"); r += 1

    CFS_R['cfads'] = r
    _lbl(ws, r, "CFADS  (EBITDA − Tax − ΔWC − Maint CapEx)", "Cr", bold=True)
    for j in range(N):
        f(r, j, f"=OPEX!{cl(j)}{OPX_R['ebitda']}"
                f"-TAX!{cl(j)}{TAX_R['tax']}"
                f"-WC!{cl(j)}{WC_R['delta']}"
                f"-CAPEX!{cl(j)}{CAP_R['maint']}", bold=True)
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR, True); r += 1

    CFS_R['debt_svc'] = r
    _lbl(ws, r, "Debt service  (interest + principal)", "Cr")
    for j in range(N):
        f(r, j, f"=DEBT!{cl(j)}{DBT_R['interest']}+DEBT!{cl(j)}{DBT_R['principal']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 1

    CFS_R['dscr'] = r
    _lbl(ws, r, "DSCR", "x", bold=True)
    for j in range(N):
        ds = CFS_R['debt_svc']
        f(r, j, f"=IF(CFS!{cl(j)}{ds}>0,MAX(CFS!{cl(j)}{CFS_R['cfads']}/CFS!{cl(j)}{ds},0),\"\")",
          FMT_MX, True)
    r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# BS — Balance Sheet (3-statement, ties to zero)
# ═══════════════════════════════════════════════════════════════════════════════
def write_bs(wb):
    ws = wb.create_sheet("BS")
    _col_widths(ws); _title_row(ws, "BALANCE SHEET", "Full 3-statement balance sheet — ties to zero every year")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, formula, fmt=FMT_CR, bold=False, fill=None):
        c = ws.cell(row=row, column=COL_YR0+j)
        c.value = formula; c.number_format = fmt; c.font = _font(bold=bold)
        if fill: c.fill = _fill(fill)

    # ── CASH ROLL-FORWARD (needed for cash & the financing gap) ──────────────
    r += 1; _hdr(ws, r, "CASH FLOW (roll-forward)"); r += 1

    BS_R['cfo'] = r
    _lbl(ws, r, "Operating cash flow (PAT + Dep − ΔWC)", "Cr")
    for j in range(N):
        f(r, j, f"=PNL!{cl(j)}{PNL_R['pat']}+DEPR!{cl(j)}{DEP_R['total']}-WC!{cl(j)}{WC_R['delta']}")
    r += 1

    BS_R['cfi'] = r
    _lbl(ws, r, "Investing cash flow (− CapEx)", "Cr")
    for j in range(N):
        f(r, j, f"=-CAPEX!{cl(j)}{CAP_R['total']}")
    r += 1

    BS_R['cff'] = r
    _lbl(ws, r, "Financing cash flow (Draw − Principal + Equity)", "Cr")
    for j in range(N):
        f(r, j, f"=DEBT!{cl(j)}{DBT_R['drawdown']}-DEBT!{cl(j)}{DBT_R['principal']}+DEBT!{cl(j)}{DBT_R['equity']}")
    r += 1

    BS_R['cash_raw'] = r
    _lbl(ws, r, "Cash before financing gap (cumulative)", "Cr")
    for j in range(N):
        chg = f"BS!{cl(j)}{BS_R['cfo']}+BS!{cl(j)}{BS_R['cfi']}+BS!{cl(j)}{BS_R['cff']}"
        if j == 0:
            f(r, j, f"={chg}")
        else:
            f(r, j, f"=BS!{cl(j-1)}{BS_R['cash_raw']}+{chg}")
    r += 1

    # ── ASSETS ───────────────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "ASSETS"); r += 1

    BS_R['nfa'] = r
    _lbl(ws, r, "Net fixed assets", "Cr")
    for j in range(N):
        f(r, j, f"=DEPR!{cl(j)}{DEP_R['nbv']}")
    r += 1

    BS_R['ar'] = r
    _lbl(ws, r, "Accounts receivable", "Cr")
    for j in range(N):
        f(r, j, f"=WC!{cl(j)}{WC_R['receivables']}")
    r += 1

    BS_R['cash'] = r
    _lbl(ws, r, "Cash & equivalents", "Cr")
    for j in range(N):
        f(r, j, f"=MAX(BS!{cl(j)}{BS_R['cash_raw']},0)")
    r += 1

    BS_R['ta'] = r
    _lbl(ws, r, "Total assets", "Cr", bold=True, fill=LTGREY)
    for j in range(N):
        f(r, j, f"=BS!{cl(j)}{BS_R['nfa']}+BS!{cl(j)}{BS_R['ar']}+BS!{cl(j)}{BS_R['cash']}", FMT_CR, True, LTGREY)
    r += 1

    # ── LIABILITIES & EQUITY ─────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "LIABILITIES & EQUITY"); r += 1

    BS_R['debt'] = r
    _lbl(ws, r, "Long-term debt", "Cr")
    for j in range(N):
        f(r, j, f"=DEBT!{cl(j)}{DBT_R['closing']}")
    r += 1

    BS_R['ap'] = r
    _lbl(ws, r, "Accounts payable", "Cr")
    for j in range(N):
        f(r, j, f"=WC!{cl(j)}{WC_R['payables']}")
    r += 1

    BS_R['addl_fin'] = r
    _lbl(ws, r, "Additional financing required", "Cr")
    for j in range(N):
        f(r, j, f"=MAX(-BS!{cl(j)}{BS_R['cash_raw']},0)")
    r += 1

    BS_R['paid_in'] = r
    _lbl(ws, r, "Paid-in equity", "Cr")
    for j in range(N):
        if j == 0:
            f(r, j, f"=DEBT!{cl(0)}{DBT_R['equity']}")
        else:
            f(r, j, f"=BS!{cl(j-1)}{BS_R['paid_in']}+DEBT!{cl(j)}{DBT_R['equity']}")
    r += 1

    BS_R['re'] = r
    _lbl(ws, r, "Retained earnings", "Cr")
    for j in range(N):
        if j == 0:
            f(r, j, f"=PNL!{cl(0)}{PNL_R['pat']}")
        else:
            f(r, j, f"=BS!{cl(j-1)}{BS_R['re']}+PNL!{cl(j)}{PNL_R['pat']}")
    r += 1

    BS_R['tle'] = r
    _lbl(ws, r, "Total liabilities & equity", "Cr", bold=True, fill=LTGREY)
    for j in range(N):
        f(r, j, f"=BS!{cl(j)}{BS_R['debt']}+BS!{cl(j)}{BS_R['ap']}+BS!{cl(j)}{BS_R['addl_fin']}"
                f"+BS!{cl(j)}{BS_R['paid_in']}+BS!{cl(j)}{BS_R['re']}", FMT_CR, True, LTGREY)
    r += 1

    BS_R['check'] = r
    _lbl(ws, r, "Balance check (Assets − L&E)", "Cr", bold=True)
    for j in range(N):
        f(r, j, f"=BS!{cl(j)}{BS_R['ta']}-BS!{cl(j)}{BS_R['tle']}", FMT_CR, True)
    r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# VAL — Valuation
# ═══════════════════════════════════════════════════════════════════════════════
def write_val(wb):
    ws = wb.create_sheet("VAL")
    _col_widths(ws); _title_row(ws, "VALUATION", "WACC, DCF, IRR — formula-driven")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def fsc(row, formula, label="", unit="", fmt=FMT_CR, bold=False, fill=None):
        """Write a scalar result in col C (not time-series)."""
        _lbl(ws, row, label, unit, bold=bold)
        c = ws.cell(row=row, column=COL_YR0)
        c.value = formula; c.number_format = fmt; c.font = _font(bold=bold)
        if fill: c.fill = _fill(fill)

    def f(row, j, formula, fmt=FMT_CR, bold=False, fill=None):
        c = ws.cell(row=row, column=COL_YR0+j)
        c.value = formula; c.number_format = fmt; c.font = _font(bold=bold)
        if fill: c.fill = _fill(fill)

    # ── WACC ──────────────────────────────────────────────────────────────
    r += 1; _hdr(ws, r, "WACC BUILD"); r += 1
    VAL_R['coe']       = r; fsc(r, f"={_asmp('cost_eq')}",  "Cost of equity",     "% p.a.", FMT_P2); r += 1
    VAL_R['cod_pre']   = r; fsc(r, f"={_asmp('int_rate')}", "Cost of debt (pre-tax)","% p.a.", FMT_P2); r += 1
    VAL_R['cod_post']  = r
    fsc(r, f"=VAL!$C${VAL_R['cod_pre']}*(1-{_asmp('tax_rate')})", "Cost of debt (post-tax)","% p.a.", FMT_P2); r += 1
    VAL_R['eq_wt']     = r; fsc(r, f"={_asmp('eq_pct')}",  "Equity weight",       "%", FMT_P2); r += 1
    VAL_R['debt_wt']   = r; fsc(r, f"={_asmp('debt_pct')}", "Debt weight",         "%", FMT_P2); r += 1
    VAL_R['wacc']      = r
    fsc(r, f"=VAL!$C${VAL_R['coe']}*VAL!$C${VAL_R['eq_wt']}"
           f"+VAL!$C${VAL_R['cod_post']}*VAL!$C${VAL_R['debt_wt']}",
        "WACC", "% p.a.", FMT_P2, True); r += 2

    # ── Terminal value ────────────────────────────────────────────────────
    _hdr(ws, r, "TERMINAL VALUE"); r += 1
    VAL_R['mult']      = r; fsc(r, f"={_asmp('ev_mult')}", "EV/EBITDA multiple",  "x", FMT_MX); r += 1
    VAL_R['ebitda10']  = r; fsc(r, f"=OPEX!L{OPX_R['ebitda']}", "EBITDA Year 10", "Cr"); r += 1
    VAL_R['tv']        = r
    fsc(r, f"=VAL!$C${VAL_R['mult']}*VAL!$C${VAL_R['ebitda10']}", "Terminal EV", "Cr", FMT_CR, True); r += 1
    VAL_R['res_debt']  = r; fsc(r, f"=DEBT!L{DBT_R['closing']}", "Residual debt (Year 10)", "Cr"); r += 1
    VAL_R['eq_tv']     = r
    fsc(r, f"=MAX(VAL!$C${VAL_R['tv']}-VAL!$C${VAL_R['res_debt']},0)",
        "Equity terminal value", "Cr", FMT_CR, True); r += 2

    # ── DCF ───────────────────────────────────────────────────────────────
    _hdr(ws, r, "DCF — PROJECT FCFF"); r += 1

    VAL_R['fcff'] = r
    _lbl(ws, r, "FCFF", "Cr")
    for j in range(N): f(r, j, f"=CFS!{cl(j)}{CFS_R['fcff']}")
    r += 1

    VAL_R['fcff_tv'] = r
    _lbl(ws, r, "FCFF incl. terminal value (last year)", "Cr", bold=True)
    for j in range(N):
        if j < N-1:
            f(r, j, f"=VAL!{cl(j)}{VAL_R['fcff']}", bold=True)
        else:
            f(r, j, f"=VAL!{cl(j)}{VAL_R['fcff']}+VAL!$C${VAL_R['tv']}", bold=True, fill=LTGREY)
    r += 1

    VAL_R['df'] = r
    _lbl(ws, r, "Discount factor  (1/(1+WACC)^yr)", "")
    for j in range(N):
        f(r, j, f"=1/(1+VAL!$C${VAL_R['wacc']})^{cl(j)}4", FMT_CR)
    r += 1

    VAL_R['pv_fcff'] = r
    _lbl(ws, r, "PV of FCFF", "Cr")
    for j in range(N):
        f(r, j, f"=VAL!{cl(j)}{VAL_R['fcff_tv']}*VAL!{cl(j)}{VAL_R['df']}")
    _w(ws, r, COL_YR0+N, f"=SUM(C{r}:L{r})", FMT_CR); r += 2

    # ── Returns summary ───────────────────────────────────────────────────
    _hdr(ws, r, "RETURNS SUMMARY"); r += 1

    VAL_R['proj_irr'] = r
    fsc(r, f"=IRR(C{VAL_R['fcff_tv']}:L{VAL_R['fcff_tv']},0.1)",
        "Project IRR", "% p.a.", FMT_P2, True); r += 1

    VAL_R['sum_pv_fcff'] = r
    fsc(r, f"=SUM(C{VAL_R['pv_fcff']}:L{VAL_R['pv_fcff']})",
        "Sum PV (FCFF, excl. TV)", "Cr"); r += 1

    VAL_R['pv_tv'] = r
    fsc(r, f"=VAL!$C${VAL_R['tv']}/(1+VAL!$C${VAL_R['wacc']})^10",
        "PV of terminal value", "Cr"); r += 1

    VAL_R['npv'] = r
    fsc(r, f"=VAL!$C${VAL_R['sum_pv_fcff']}+VAL!$C${VAL_R['pv_tv']}",
        "NPV (at WACC)", "Cr", FMT_CR, True); r += 2

    # ── Equity returns ────────────────────────────────────────────────────
    _hdr(ws, r, "EQUITY RETURNS"); r += 1

    VAL_R['fcfe'] = r
    _lbl(ws, r, "FCFE", "Cr")
    for j in range(N): f(r, j, f"=CFS!{cl(j)}{CFS_R['fcfe']}")
    r += 1

    VAL_R['fcfe_tv'] = r
    _lbl(ws, r, "FCFE incl. equity terminal value", "Cr", bold=True)
    for j in range(N):
        if j < N-1:
            f(r, j, f"=VAL!{cl(j)}{VAL_R['fcfe']}", bold=True)
        else:
            f(r, j, f"=VAL!{cl(j)}{VAL_R['fcfe']}+VAL!$C${VAL_R['eq_tv']}", bold=True, fill=LTGREY)
    r += 1

    VAL_R['eq_irr'] = r
    fsc(r, f"=IRR(C{VAL_R['fcfe_tv']}:L{VAL_R['fcfe_tv']},0.1)",
        "Equity IRR", "% p.a.", FMT_P2, True); r += 1

    VAL_R['eq_inv'] = r
    fsc(r, f"=SUM(CAPEX!C{CAP_R['eq_fund']}:L{CAP_R['eq_fund']})",
        "Total equity invested", "Cr"); r += 1

    VAL_R['moic'] = r
    fsc(r, f"=SUMIF(C{VAL_R['fcfe_tv']}:L{VAL_R['fcfe_tv']},\">0\")/VAL!$C${VAL_R['eq_inv']}",
        "MOIC (equity multiple)", "x", FMT_MX, True); r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# DASH — Executive dashboard
# ═══════════════════════════════════════════════════════════════════════════════
def write_dash(wb):
    ws = wb.create_sheet("DASH")
    _col_widths(ws); _title_row(ws, "EXECUTIVE DASHBOARD",
                                 "Greenfield Data Center | Mumbai | 1,000 Racks | 2026–2035")
    ws.freeze_panes = "C5"

    r = 3; _yr_hdrs(ws, r, r+1); r += 2

    def f(row, j, formula, fmt=FMT_CR, bold=False, fill=None):
        c = ws.cell(row=row, column=COL_YR0+j)
        c.value = formula; c.number_format = fmt; c.font = _font(bold=bold)
        if fill: c.fill = _fill(fill)

    r += 1; _hdr(ws, r, "KEY METRICS AT A GLANCE"); r += 1

    metrics = [
        ("Total CapEx",          f"=SUM(CAPEX!C{CAP_R['total']}:L{CAP_R['total']})", "Cr",   FMT_CR),
        ("NPV (at WACC)",        f"=VAL!$C${VAL_R['npv']}",                    "Cr",   FMT_CR),
        ("Project IRR",          f"=VAL!$C${VAL_R['proj_irr']}",               "% pa", FMT_P2),
        ("Equity IRR",           f"=VAL!$C${VAL_R['eq_irr']}",                 "% pa", FMT_P2),
        ("MOIC",                 f"=VAL!$C${VAL_R['moic']}",                   "x",    FMT_MX),
        ("EBITDA Yr 10",         f"=OPEX!L{OPX_R['ebitda']}",                  "Cr",   FMT_CR),
        ("EBITDA Margin Yr 10",  f"=OPEX!L{OPX_R['ebitda_m']}",                "%",    FMT_P1),
        ("Terminal Value (EV)",  f"=VAL!$C${VAL_R['tv']}",                     "Cr",   FMT_CR),
    ]
    for label, fml, unit, fmt in metrics:
        _lbl(ws, r, label, unit, bold=True)
        _w(ws, r, COL_YR0, fml, fmt, bold=True, fill=LTGREY, ah="right"); r += 1

    r += 1; _hdr(ws, r, "REVENUE SUMMARY"); r += 1
    _lbl(ws, r, "Net revenue", "Cr")
    for j in range(N): f(r, j, f"=REV!{cl(j)}{REV_R['net']}")
    r += 1

    _lbl(ws, r, "EBITDA", "Cr", bold=True)
    for j in range(N): f(r, j, f"=OPEX!{cl(j)}{OPX_R['ebitda']}", bold=True)
    r += 1

    _lbl(ws, r, "EBITDA margin", "%")
    for j in range(N): f(r, j, f"=OPEX!{cl(j)}{OPX_R['ebitda_m']}", FMT_P1)
    r += 1

    _lbl(ws, r, "PAT", "Cr")
    for j in range(N): f(r, j, f"=PNL!{cl(j)}{PNL_R['pat']}")
    r += 2

    _hdr(ws, r, "DEBT SERVICE"); r += 1
    _lbl(ws, r, "Closing debt", "Cr")
    for j in range(N): f(r, j, f"=DEBT!{cl(j)}{DBT_R['closing']}")
    r += 1

    _lbl(ws, r, "CFADS", "Cr")
    for j in range(N): f(r, j, f"=CFS!{cl(j)}{CFS_R['cfads']}")
    r += 1

    _lbl(ws, r, "DSCR", "x", bold=True)
    for j in range(N): f(r, j, f"=CFS!{cl(j)}{CFS_R['dscr']}", FMT_MX, True)
    r += 1

    _col_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# COVER — Navigation
# ═══════════════════════════════════════════════════════════════════════════════
def write_cover(wb):
    ws = wb.create_sheet("COVER", 0)  # insert at position 0
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20

    ws.merge_cells("A1:C2")
    tc = ws.cell(1, 1)
    tc.value = "  Greenfield Data Center — Financial Model"
    tc.font  = _font(True, WHITE, 14); tc.fill = _fill(NAVY); tc.alignment = _aln("left")
    ws.row_dimensions[1].height = 30; ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:C3")
    sc = ws.cell(3, 1)
    sc.value = f"  Mumbai | 1,000 Racks | {YEARS[0]}–{YEARS[-1]}"
    sc.font  = _font(ital=True, col=MDGREY); sc.alignment = _aln("left")

    sheets = [
        ("ASMP",  "All assumptions — change here only"),
        ("SIZE",  "Rack deployment, occupancy & power"),
        ("REV",   "Revenue build"),
        ("CAPEX", "Capital expenditure schedule"),
        ("OPEX",  "OpEx & EBITDA"),
        ("DEPR",  "Book depreciation (SLM)"),
        ("DEBT",  "Loan schedule & tranche detail"),
        ("TAX",   "Taxation (simplified)"),
        ("WC",    "Working capital"),
        ("PNL",   "Profit & Loss statement"),
        ("CFS",   "Cash flow statement — FCFF, FCFE, DSCR"),
        ("BS",    "Balance sheet — assets, liabilities & equity"),
        ("VAL",   "Valuation — WACC, DCF, IRR"),
        ("DASH",  "Executive dashboard"),
    ]

    r = 5
    ws.cell(r, 1).value = "Sheet"; ws.cell(r, 1).font = _font(True, WHITE)
    ws.cell(r, 1).fill = _fill(NAVY); ws.cell(r, 1).alignment = _aln("left")
    ws.cell(r, 2).value = "Description"; ws.cell(r, 2).font = _font(True, WHITE)
    ws.cell(r, 2).fill = _fill(NAVY); ws.cell(r, 2).alignment = _aln("left")
    r += 1

    for name, desc in sheets:
        c = ws.cell(r, 1)
        c.value = name; c.hyperlink = f"#{name}!A1"
        c.font  = Font(name="Calibri", color="1F3864", underline="single", bold=True)
        c.alignment = _aln("left")
        d = ws.cell(r, 2)
        d.value = desc; d.font = _font(col=DKGREY); d.alignment = _aln("left")
        if r % 2 == 0:
            for col in [1, 2, 3]:
                ws.cell(r, col).fill = _fill(LTGREY)
        r += 1

    r += 1
    ws.cell(r, 1).value = "Colour coding:"
    ws.cell(r, 1).font  = _font(True, DKGREY)
    r += 1
    legend = [
        (BLUE_IN, "Blue   — hardcoded input (change only in ASMP)"),
        (SRC_IN,  "Teal   — market-sourced (AI) input — rack MRC, tariff, markup, PUE, interest, land, civil/elec/mech"),
        (LTGREY,  "Grey   — calculated total"),
        (GRNCELL, "Green  — cross-sheet formula"),
        ("FFF9E6", "Amber  — derived assumption (formula in ASMP)"),
    ]
    for fill_hex, label in legend:
        ws.cell(r, 1).fill = _fill(fill_hex)
        ws.cell(r, 2).value = label; ws.cell(r, 2).font = _font(col=DKGREY)
        ws.cell(r, 2).alignment = _aln("left")
        r += 1

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def _predeclare_rows():
    """Pre-declare row numbers for sheets that have circular cross-references.
    PNL references TAX rows; TAX references PNL rows — both must be known before either is written.
    Row numbers are determined by the fixed layout in each sheet writer (r starts at 3, +2 for
    year headers, +1 for blank, +1 for section header, then one row per line item).
    """
    # PNL layout: header at row 6, data from row 7
    PNL_R.update({
        'net_rev': 7, 'total_opx': 8, 'ebitda': 9, 'ebitda_m': 10,
        'dep': 11, 'ebit': 12, 'interest': 13, 'ebt': 14,
        'tax': 15, 'pat': 16, 'pat_m': 17,
    })
    # TAX layout: header at row 6, data from row 7
    TAX_R.update({'ebt': 7, 'cumul_ebt': 8, 'tax': 9})
    # WC layout (for CFS forward ref safety)
    WC_R.update({'nwc': 7, 'delta': 8})


def generate(out_path: str = "outputs/excel_models/dcf_model.xlsx", override: dict = None):
    """Build the workbook. If `override` is provided (from an API run), the
    workbook reflects that run's inputs and assumptions; otherwise it uses
    the module defaults. `override` keys: ui, rev, cap, opx, dep, loan, tax,
    wc, val."""
    global USER_INPUTS, N, YEARS, _OVERRIDE
    _OVERRIDE = override
    # Deterministically set module state every call so a prior run's params
    # never leak into a later default generation.
    USER_INPUTS = override["ui"] if (override and override.get("ui")) else dict(_DEFAULT_USER_INPUTS)
    N = USER_INPUTS["projection_years"]
    YEARS = list(range(USER_INPUTS["start_year"], USER_INPUTS["start_year"] + N))

    print("Running pipeline…")
    P = _run_pipeline()

    # Pre-declare row registries for cross-sheet forward references
    _predeclare_rows()

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    print("Writing ASMP…")
    write_asmp(wb, P)

    print("Writing SIZE…")
    write_size(wb)

    print("Writing REV…")
    write_rev(wb)

    print("Writing CAPEX…")
    write_capex(wb)

    print("Writing OPEX…")
    write_opex(wb)

    print("Writing DEPR…")
    write_depr(wb)

    # DEBT must come before TAX/PNL (PNL refs DEBT.interest)
    print("Writing DEBT…")
    write_debt(wb)

    # PNL must be written before TAX (TAX refs PNL.ebt)
    print("Writing PNL…")
    write_pnl(wb)

    print("Writing TAX…")
    write_tax(wb)

    print("Writing WC…")
    write_wc(wb)

    print("Writing CFS…")
    write_cfs(wb)

    print("Writing BS…")
    write_bs(wb)

    print("Writing VAL…")
    write_val(wb)

    print("Writing DASH…")
    write_dash(wb)

    print("Writing COVER…")
    write_cover(wb)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path


if __name__ == "__main__":
    generate()
